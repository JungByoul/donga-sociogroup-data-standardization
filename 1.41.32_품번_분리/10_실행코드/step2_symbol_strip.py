# -*- coding: utf-8 -*-
r"""
v1.1 -> v1.2 (2차 품목분리)
품번_코어가 아직 비어있는 나머지 278건 중, 대괄호([42609])나 # 접두어(#0591)처럼
장식기호만 붙어 실제 값 판독에 지장 없는 6건을 찾아 기호를 제거한 값을
기존 품번_코어 열에 채운다. (단위분리와 같은 "핵심코드 정제" 개념 재사용)
"""
import re
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.32_품번_분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.1.1.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.32_품번_분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.1.2.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_Q_CAS원본 = 17          # Q
COL_CAS_3차판정 = 21        # U
COL_품번_코어 = 22          # V
LABEL_ITEM_NO = "[품번/문서번호]"

BRACKET_RE = re.compile(r"^\[(.+)\]$")
HASH_RE = re.compile(r"^#(.+)$")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    stripped = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_CAS_3차판정).value != LABEL_ITEM_NO:
            continue
        core_cell = ws.cell(row=r, column=COL_품번_코어)
        if core_cell.value not in (None, ""):
            continue  # 이미 단위분리 등으로 처리된 행은 건드리지 않음

        q = ws.cell(row=r, column=COL_Q_CAS원본).value
        q_str = "" if q is None else str(q).strip()

        m = BRACKET_RE.match(q_str) or HASH_RE.match(q_str)
        if m:
            core_cell.value = m.group(1)
            stripped += 1

    wb.save(DST)
    print(f"[요약] 장식기호 제거: {stripped}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
