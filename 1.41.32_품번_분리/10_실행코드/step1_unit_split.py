# -*- coding: utf-8 -*-
r"""
v0.7 -> v1.1
CAS_3차판정이 [품번/문서번호]인 289건 중, "코드-숫자단위"(예: DUO82049-20L) 형태로
대시 뒤에 숫자+단위(ML/MG/UG/KG/RXN/G/L)가 명확히 붙어있는 값만 코어 품번과 단위로 분리.
붙어있는 형태(예: 9272S)나 애매한 건 건드리지 않음 - 사용자 확인 후 확정한 범위(A유형만).
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\40_중간수정\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.7.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.32_품번_분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.1.1.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_Q_CAS원본 = 17          # Q
COL_CAS_3차판정 = 21        # U
LABEL_ITEM_NO = "[품번/문서번호]"

INSERT_AT = 22             # U 다음 (기존 공급사명 앞)
N_NEW_COLS = 2

UNIT_RE = re.compile(r"^(.+)-(\d+)\s*(ML|MG|UG|KG|RXN|G|L)$", re.IGNORECASE)

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    computed = {}  # row -> (core, unit)
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_CAS_3차판정).value != LABEL_ITEM_NO:
            continue
        q = ws.cell(row=r, column=COL_Q_CAS원본).value
        q_str = "" if q is None else str(q).strip()
        m = UNIT_RE.match(q_str)
        if m:
            computed[r] = (m.group(1), (m.group(2) + m.group(3)).upper())

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    ws.cell(row=4, column=INSERT_AT, value="품번_코어")
    ws.cell(row=4, column=INSERT_AT + 1, value="품번_단위")

    for r, (core, unit) in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=core)
        ws.cell(row=r, column=INSERT_AT + 1, value=unit)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] 단위 분리(A유형): {len(computed)}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
