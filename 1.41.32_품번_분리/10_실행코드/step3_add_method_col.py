# -*- coding: utf-8 -*-
r"""
v1.2 -> v1.3
품번_코어는 있는데 품번_단위가 비어있으면 "단위분리 안 된 건지 다른 이유로 정제된 건지"
구분이 안 됨 -> "정제방식" 열을 추가해서 각 행에 어떤 처리를 했는지 명시.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.32_품번_분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.1.2.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.32_품번_분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.1.3.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_품번_코어 = 22   # V
COL_품번_단위 = 23   # W
INSERT_AT = 24       # W 다음

METHOD_UNIT_SPLIT = "단위분리"
METHOD_SYMBOL_STRIP = "기호제거"

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    computed = {}
    n_unit = n_symbol = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        core = ws.cell(row=r, column=COL_품번_코어).value
        unit = ws.cell(row=r, column=COL_품번_단위).value
        if core in (None, ""):
            continue
        if unit not in (None, ""):
            computed[r] = METHOD_UNIT_SPLIT
            n_unit += 1
        else:
            computed[r] = METHOD_SYMBOL_STRIP
            n_symbol += 1

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, 1)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, 1)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    ws.cell(row=4, column=INSERT_AT, value="정제방식")
    for r, method in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=method)

    wb.save(DST)
    print(f"[요약] 단위분리: {n_unit}건 / 기호제거: {n_symbol}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
