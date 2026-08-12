# -*- coding: utf-8 -*-
r"""
O열(품번) 1차 판정
O열 바로 뒤에 "O열_판정" 열을 추가.
- 공란: 그대로 공란
- 값없음 플레이스홀더('-', 한/두 글자, '해당없음'류): [값없음]
- CAS 형태+체크디지트 통과: [CAS의심] (Q열/CAS No.(수정)은 건드리지 않음)
- 설명문/URL 등 명백한 비품번(키워드 매칭): [비품번]
- 나머지: [품번]
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.33_CAS_품번_분리_최종(Q열)\40_송부용\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_최종.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.0.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_O_품번 = 15
INSERT_AT = 16  # O 다음

_CAS_RE = re.compile(r"\b(\d{2,7}-\d{2}-\d)\b")
PLACEHOLDER_EXACT = {"-", "해당없음", "품번따로없음", "없음", "n/a", "na"}
DESC_KEYWORDS = [
    "mouse", "female", "male", "system", "server", "license", "standard",
    "core", "ref.", "type ", "week", "cage", "centrifuge", "triple", "quad",
    "agilent", "sciex", "windows", "주령", "시험 중",
]

LABEL_NONE = "[값없음]"
LABEL_CAS = "[CAS의심]"
LABEL_DESC = "[비품번]"
LABEL_ITEM = "[품번]"

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def cas_checksum_valid(cas: str) -> bool:
    digits = cas.replace("-", "")
    body, check = digits[:-1], int(digits[-1])
    total = sum(int(d) * (i + 1) for i, d in enumerate(reversed(body)))
    return total % 10 == check


def classify(s: str) -> str:
    s_lower = s.lower()
    if s_lower in PLACEHOLDER_EXACT or len(s) <= 2:
        return LABEL_NONE
    m = _CAS_RE.search(s)
    if m and cas_checksum_valid(m.group(1)):
        return LABEL_CAS
    if "http" in s_lower or any(k in s_lower for k in DESC_KEYWORDS):
        return LABEL_DESC
    return LABEL_ITEM


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    computed = {}
    counts = {LABEL_NONE: 0, LABEL_CAS: 0, LABEL_DESC: 0, LABEL_ITEM: 0}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(row=r, column=COL_O_품번).value
        if v in (None, ""):
            continue
        label = classify(str(v).strip())
        computed[r] = label
        counts[label] += 1

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, 1)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, 1)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    letter = get_column_letter(INSERT_AT)
    ws.column_dimensions[letter].hidden = False
    ws.column_dimensions[letter].width = 14

    ws.cell(row=4, column=INSERT_AT, value="O열_판정")
    for r, label in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=label)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] {counts}")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
