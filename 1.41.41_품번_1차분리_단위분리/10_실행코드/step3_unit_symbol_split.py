# -*- coding: utf-8 -*-
r"""
v1.0 -> v1.1
O열_판정이 [품번]인 13,964건 중, Q열 파이프라인에서 썼던 것과 동일한 규칙으로
단위분리(A유형: 코드-숫자단위)와 기호제거(대괄호/#)를 적용.
Q열 쪽 컬럼(품번_코어 등)과 이름이 겹치지 않도록 O열_ 접두어를 붙인 별도 컬럼 사용.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.0.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.1.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_O_품번 = 15
COL_O열_판정 = 16
LABEL_ITEM = "[품번]"

INSERT_AT = 17  # O열_판정(16) 다음
N_NEW_COLS = 3

UNIT_RE = re.compile(r"^(.+)-(\d+)\s*(ML|MG|UG|KG|RXN|G|L)$", re.IGNORECASE)
BRACKET_RE = re.compile(r"^\[(.+)\]$")
HASH_RE = re.compile(r"^#(.+)$")

METHOD_UNIT_SPLIT = "단위분리"
METHOD_SYMBOL_STRIP = "기호제거"

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    computed = {}  # row -> (core, unit, method)
    n_unit = n_symbol = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_O열_판정).value != LABEL_ITEM:
            continue
        v = ws.cell(row=r, column=COL_O_품번).value
        s = "" if v is None else str(v).strip()

        m = UNIT_RE.match(s)
        if m:
            core = m.group(1)
            unit = (m.group(2) + m.group(3)).upper()
            computed[r] = (core, unit, METHOD_UNIT_SPLIT)
            n_unit += 1
            continue

        m = BRACKET_RE.match(s) or HASH_RE.match(s)
        if m:
            computed[r] = (m.group(1), "", METHOD_SYMBOL_STRIP)
            n_symbol += 1

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    for i in range(N_NEW_COLS):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].hidden = False
        ws.column_dimensions[letter].width = 16

    ws.cell(row=4, column=INSERT_AT, value="O열_품번_코어")
    ws.cell(row=4, column=INSERT_AT + 1, value="O열_품번_단위")
    ws.cell(row=4, column=INSERT_AT + 2, value="O열_정제방식")

    for r, (core, unit, method) in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=core)
        ws.cell(row=r, column=INSERT_AT + 1, value=unit)
        ws.cell(row=r, column=INSERT_AT + 2, value=method)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] 단위분리: {n_unit}건 / 기호제거: {n_symbol}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
