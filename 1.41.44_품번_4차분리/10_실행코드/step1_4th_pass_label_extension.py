# -*- coding: utf-8 -*-
r"""
v2.1 -> v2.2

v2.1은 1.41.43.5(1~3차 작업 개편)에서 만든 v1.9를 그대로 복사한 것으로, O열_4차판정에
이미 [코드뭉침]/[설명문혼입]/[단위값]/[구조식의심](31건) + O열_4차_검토메모(슬래시 151건)가
반영돼 있음. 여기서는 그 위에 추가로, 아직 라벨/메모가 없는 값 중 아래 두 패턴을 [설명문혼입]로
추가 확정한다(기존 [설명문혼입] 규칙과 같은 성격 — "설명 라벨이 품번 자리에 섞여 들어간 값"의
탐지 범위를 넓힌 것뿐이라 새 라벨을 만들지 않고 그대로 재사용).

추가 규칙(둘 다 오탐 거의 없음을 샘플 확인함):
1) 콜론(:) 포함 -> "Model: 22.046", "S/N : 46106"처럼 "라벨: 값" 형식
2) 한글 라벨 키워드 포함 -> 견적번호/견적서번호/일련번호/시리얼번호/부품번호/모델명

전체 13,669건 풀 중 순수숫자만(3,632건, Sigma 등 실제 카탈로그 번호와 겹쳐 위험)이나
제조사별 형태이상치(제조사 내 지배형태 비중이 8~50%로 낮아 신뢰 불가)는 검증 결과 자동라벨링
근거로 쓰기엔 오탐 위험이 커서 이번 4차판정에는 포함하지 않음.
"""
import re
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v2.1.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v2.2.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5

COL_품번 = 15
COL_1차판정 = 16
COL_2차판정 = 17
COL_3차판정 = 18
COL_4차판정 = 19
COL_4차_검토메모 = 20
COL_품번_코어 = 21

LABEL_ITEM = "[품번]"
LABEL_DESC = "[설명문혼입]"

_LABEL_KW_RE = re.compile(r"견적\s*(서)?\s*번호|일련번호|시리얼\s*번호|부품\s*번호|모델명")


def is_desc_leak(v: str) -> bool:
    return (":" in v) or bool(_LABEL_KW_RE.search(v))


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n_added = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_1차판정).value != LABEL_ITEM:
            continue
        if ws.cell(row=r, column=COL_2차판정).value or ws.cell(row=r, column=COL_3차판정).value:
            continue
        if ws.cell(row=r, column=COL_4차판정).value or ws.cell(row=r, column=COL_4차_검토메모).value:
            continue

        core = ws.cell(row=r, column=COL_품번_코어).value
        raw = ws.cell(row=r, column=COL_품번).value
        final_value = str(core).strip() if core not in (None, "") else str(raw).strip()

        if is_desc_leak(final_value):
            ws.cell(row=r, column=COL_4차판정, value=LABEL_DESC)
            n_added += 1

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] O열_4차판정 [설명문혼입] 추가(콜론/한글라벨): {n_added}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
