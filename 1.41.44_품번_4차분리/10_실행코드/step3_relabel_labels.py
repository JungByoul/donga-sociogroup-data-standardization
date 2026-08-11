# -*- coding: utf-8 -*-
r"""
v1.8 -> v1.9

O열_1차판정과 O열_3차판정이 서로 다른 기준인데도 똑같이 "[비품번]" 텍스트를 써서
헷갈리기 쉬웠던 문제를 해결. 열 구조/판정 로직은 그대로 두고 라벨 텍스트만 구분되게 변경.

- O열_1차판정: [비품번] -> [설명문키워드]  (mouse/system/server/standard/sciex 등
  고정 키워드 리스트 + http 포함 여부로 판정된 것. 1.41.41 폴더의 step2_build_final.py 참고)
- O열_3차판정: [비품번] -> [숫자없음]      (값에 숫자가 하나도 없다는 규칙으로 판정된 것.
  같은 폴더의 step2_2nd_3rd_rebuild_and_4th_reapply.py 참고)
"""
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v1.8.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v1.9.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_1차판정 = 16
COL_3차판정 = 18

OLD_LABEL = "[비품번]"
NEW_LABEL_1차 = "[설명문키워드]"
NEW_LABEL_3차 = "[숫자없음]"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n1 = n3 = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        c1 = ws.cell(row=r, column=COL_1차판정)
        if c1.value == OLD_LABEL:
            c1.value = NEW_LABEL_1차
            n1 += 1
        c3 = ws.cell(row=r, column=COL_3차판정)
        if c3.value == OLD_LABEL:
            c3.value = NEW_LABEL_3차
            n3 += 1

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] O열_1차판정 [비품번] -> {NEW_LABEL_1차}: {n1}건")
    print(f"[요약] O열_3차판정 [비품번] -> {NEW_LABEL_3차}: {n3}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
