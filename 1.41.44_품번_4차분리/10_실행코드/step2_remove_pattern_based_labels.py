# -*- coding: utf-8 -*-
r"""
v2.2 -> v2.3

사용자 지시: [코드뭉침](쉼표 포함, 20건)과 [설명문혼입](for/Mfr#/콜론/한글라벨 키워드, 24건)은
정규식/키워드 패턴 기반 자동판정이라 신뢰도가 낮으므로 라벨 자체를 폐기한다.
[단위값](5건)/[구조식의심](1건)은 육안 확인 후 값을 하드코딩(정확일치)한 것이라 그대로 유지.

슬래시(/) 포함 151건 검토메모("슬래시포함(수동검토필요)")도 같은 이유(패턴 기반)로 함께 제거한다.
이 메모는 애초에 확정 라벨이 아니라 "사람이 봐야 할 후보" 표시였음.

해당 44건(라벨) + 151건(메모)은 O열_4차판정/O열_4차_검토메모를 다시 빈 값으로 되돌려
라벨 없음(=아직 품번 후보로 남음) 상태로 복귀시킨다.
"""
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v2.2.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v2.3.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5

COL_4차판정 = 19
COL_4차_검토메모 = 20

LABELS_TO_DROP = {"[코드뭉침]", "[설명문혼입]"}
NOTE_TO_DROP = "슬래시포함(수동검토필요)"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n_removed = {label: 0 for label in LABELS_TO_DROP}
    n_note_removed = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        cell = ws.cell(row=r, column=COL_4차판정)
        if cell.value in LABELS_TO_DROP:
            n_removed[cell.value] += 1
            cell.value = None

        memo_cell = ws.cell(row=r, column=COL_4차_검토메모)
        if memo_cell.value == NOTE_TO_DROP:
            n_note_removed += 1
            memo_cell.value = None

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print("[요약] O열_4차판정 라벨 폐기(빈 값으로 복귀)")
    for label, n in n_removed.items():
        print(f"  {label}: {n}건")
    print(f"  라벨 합계: {sum(n_removed.values())}건")
    print(f"[요약] O열_4차_검토메모({NOTE_TO_DROP}) 폐기: {n_note_removed}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
