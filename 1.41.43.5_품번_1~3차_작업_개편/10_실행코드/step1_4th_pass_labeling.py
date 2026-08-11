# -*- coding: utf-8 -*-
r"""
v1.6 -> v1.7

[주의] 이 v1.7 산출물은 이후 2차/3차판정 기준이 바뀌면서 step2가 v1.6에서 처음부터
다시 만든 v1.8로 대체됨. v1.7은 실제로 안 쓰이는 중간 산출물이지만 작업 이력 보존 차원에서 남겨둠.

3차판정에서 남은 '품번(잠정)' 13,678건에 대해 패턴 분석(260810_품번_패턴분석_결과.xlsx)에서
원문을 직접 대조 확인한 결과를 바탕으로 O열_4차판정 신규 열을 추가한다.

라벨 규칙(전수 적용, 확실한 케이스만 자동 라벨링):
1) [코드뭉침]   : 최종값에 쉼표(,) 포함 -> 여러 코드/설명이 한 칸에 나열된 것으로 20건 전부 육안 확인함
2) [설명문혼입] : 콤마 규칙에 안 걸리면서 "for"(단어경계) 또는 "Mfr#"/"Item#" 패턴 포함
                 -> "cable for X&Y", "... for CV60-35-20 sensor" 등 설명문이 그대로 들어간 경우
3) [단위값]     : 육안 확인으로 실제 코드 없이 구매단위만 들어있음을 확정한 값 5건(하드코딩, 정확일치)
                 -> 정규식(숫자+단위)으로 자동판정하면 Cell Signaling/Terumo/GASTEC 등
                    정상 카탈로그 코드(2532L, 32506M, 11L 등)까지 오탐되어 하드코딩 화이트리스트로 처리
4) [구조식의심] : SMILES 화학구조식으로 확인된 값 1건(하드코딩, 정확일치)

위 4개 규칙에 안 걸리지만 슬래시(/)를 포함한 나머지 151건은 자동판정하지 않는다.
(표본 확인 결과 대부분 정상적인 단일 코드 표기 방식이라 규칙화 시 오탐 위험이 큼)
-> O열_4차_검토메모 열에 "슬래시포함(수동검토필요)"만 표시해 사람이 개별 판단하도록 남겨둔다.

'최종값' 재구성 로직은 3차 분리 패턴분석(step2)과 동일: O열_품번_코어가 있으면 그 값, 없으면 원본 품번 값.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43_품번_3차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번3차판정_v1.6.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43.5_품번_1~3차_작업_개편\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v1.7.xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_품번 = 15
COL_1차판정 = 16
COL_2차판정 = 17
COL_3차판정 = 18
COL_품번_코어 = 19
COL_제조사 = 22

INSERT_AT = 19  # O열_3차판정(18) 다음, O열_품번_코어(19) 앞
N_NEW_COLS = 2  # O열_4차판정 + O열_4차_검토메모

LABEL_ITEM = "[품번]"
LABEL_BUNDLE = "[코드뭉침]"
LABEL_DESC = "[설명문혼입]"
LABEL_UNIT = "[단위값]"
LABEL_STRUCT = "[구조식의심]"
NOTE_SLASH = "슬래시포함(수동검토필요)"

UNIT_ONLY_WHITELIST = {"50g", "25mg", "50mg", "500mg", "8mm"}
STRUCT_WHITELIST = {"OC1(CCOS(C2=CC=C(C)C=C2)(=O)=O)CCN(C(OCC3=CC=CC=C3"}

_FOR_RE = re.compile(r"\bfor\b", re.I)
_MFR_RE = re.compile(r"Mfr\s*#|Item\s*#", re.I)
_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def classify(final_value: str) -> tuple[str, str]:
    v = final_value
    if "," in v:
        return LABEL_BUNDLE, ""
    if _FOR_RE.search(v) or _MFR_RE.search(v):
        return LABEL_DESC, ""
    if v in UNIT_ONLY_WHITELIST:
        return LABEL_UNIT, ""
    if v in STRUCT_WHITELIST:
        return LABEL_STRUCT, ""
    if "/" in v:
        return "", NOTE_SLASH
    return "", ""


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    computed = {}  # row -> (label, note)
    counts = {LABEL_BUNDLE: 0, LABEL_DESC: 0, LABEL_UNIT: 0, LABEL_STRUCT: 0}
    n_note = 0

    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_1차판정).value != LABEL_ITEM:
            continue
        if ws.cell(row=r, column=COL_2차판정).value or ws.cell(row=r, column=COL_3차판정).value:
            continue

        core = ws.cell(row=r, column=COL_품번_코어).value
        raw = ws.cell(row=r, column=COL_품번).value
        final_value = str(core).strip() if core not in (None, "") else str(raw).strip()

        label, note = classify(final_value)
        if label or note:
            computed[r] = (label, note)
            if label:
                counts[label] += 1
            if note:
                n_note += 1

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    for i in range(N_NEW_COLS):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].hidden = False
        ws.column_dimensions[letter].width = 22

    ws.cell(row=HEADER_ROW, column=INSERT_AT, value="O열_4차판정")
    ws.cell(row=HEADER_ROW, column=INSERT_AT + 1, value="O열_4차_검토메모")

    for r, (label, note) in computed.items():
        if label:
            ws.cell(row=r, column=INSERT_AT, value=label)
        if note:
            ws.cell(row=r, column=INSERT_AT + 1, value=note)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print("[요약] O열_4차판정 자동 라벨링 결과 (3차판정에서 남은 13,678건 대상)")
    for k, c in counts.items():
        print(f"  {k}: {c}건")
    print(f"  검토메모({NOTE_SLASH}): {n_note}건")
    print(f"  라벨/메모 없음(그대로 품번 후보 유지): {13678 - sum(counts.values()) - n_note}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
