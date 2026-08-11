# -*- coding: utf-8 -*-
r"""
v1.3 -> 최종본
Q열과 R열(CAS_1차판정) 사이에 "품번(수정)" / "CAS No.(수정)" 2개 열을 추가.

CAS No.(수정):
  - 2차 보정값(CAS_2차판정)이 있으면 그 값
  - 없고 1차판정이 유효 CAS면(즉 [CAS아님]이 아니면) 1차판정 값
  - 둘 다 없으면(품번/미상/화학식 등 CAS로 확정되지 않은 행) 원본 Q열 값 그대로

품번(수정):
  - CAS_3차판정이 [품번/문서번호]인 행만 대상
  - 품번_코어가 있으면 그 값, 없으면 원본 Q열 값 그대로
  - 품번/문서번호가 아닌 행은 공란
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.32_품번_분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.1.3.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.33_CAS_품번_분리_최종\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_최종.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5

COL_Q = 17               # Q (CAS No. 원본)
COL_CAS1 = 18             # CAS_1차판정
COL_CAS2 = 20             # CAS_2차판정
COL_CAS3 = 21             # CAS_3차판정
COL_품번코어 = 22          # 품번_코어

INSERT_AT = 18            # Q 다음, CAS_1차판정 앞
N_NEW_COLS = 2
LABEL_ITEM_NO = "[품번/문서번호]"
LABEL_CAS_NONE = "[CAS아님]"

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    computed = {}  # row -> (품번수정, CASNo수정)
    n_cas_from_2 = n_cas_from_1 = n_cas_fallback = 0
    n_item_core = n_item_fallback = 0

    for r in range(DATA_START_ROW, ws.max_row + 1):
        q = ws.cell(row=r, column=COL_Q).value
        cas1 = ws.cell(row=r, column=COL_CAS1).value
        cas2 = ws.cell(row=r, column=COL_CAS2).value
        cas3 = ws.cell(row=r, column=COL_CAS3).value
        core = ws.cell(row=r, column=COL_품번코어).value

        if cas2 not in (None, ""):
            cas_final = cas2
            n_cas_from_2 += 1
        elif cas1 not in (None, "", LABEL_CAS_NONE):
            cas_final = cas1
            n_cas_from_1 += 1
        elif q not in (None, ""):
            cas_final = q
            n_cas_fallback += 1
        else:
            cas_final = None

        item_final = None
        if cas3 == LABEL_ITEM_NO:
            if core not in (None, ""):
                item_final = core
                n_item_core += 1
            else:
                item_final = q
                n_item_fallback += 1

        if cas_final is not None or item_final is not None:
            computed[r] = (item_final, cas_final)

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    ws.cell(row=4, column=INSERT_AT, value="품번(수정)")
    ws.cell(row=4, column=INSERT_AT + 1, value="CAS No.(수정)")

    # openpyxl insert_cols는 column_dimensions(너비/숨김)를 제대로 안 옮겨서
    # 새 열이 우연히 원래 숨김 처리돼 있던 문자 위치를 물려받을 수 있음 - 명시적으로 해제
    for letter in (get_column_letter(INSERT_AT), get_column_letter(INSERT_AT + 1)):
        ws.column_dimensions[letter].hidden = False
        ws.column_dimensions[letter].width = 16

    for r, (item_final, cas_final) in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=item_final)
        ws.cell(row=r, column=INSERT_AT + 1, value=cas_final)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] CAS No.(수정) - 2차보정: {n_cas_from_2} / 1차유효: {n_cas_from_1} / 원본유지(품번등): {n_cas_fallback}")
    print(f"[요약] 품번(수정) - 코어값: {n_item_core} / 원본유지: {n_item_fallback}")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
