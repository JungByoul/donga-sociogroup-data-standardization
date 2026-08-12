# -*- coding: utf-8 -*-
r"""
v3.1: v3.0에서 회사명으로 게이트돼있던 '제거형' 규칙들을 전부 전역(내용 기반) 규칙으로 전환.

바뀐 것:
- 라벨 프리픽스/서픽스 제거(견적번호/부품 번호/USP/시리얼 번호/모델명/끝 외)를 특정 회사명 조건
  없이, 값이 그 패턴에 맞으면 어느 회사든 적용.
- 선두 '#' 제거를 전 회사 공통 적용(기존: 대한과학/Cell Signaling/Sigma/Sigma-Aldrich 전용).
- 전체가 '[...]'로 감싸진 경우 제거를 전 회사 공통 적용(기존: Sartorius 전용).
- 끝 'UI' 제거를 전 회사 공통 적용(기존: Agilent 전용).

그대로 유지한 것(사용자 확인 완료 — 회사별 안전장치라 무조건 전역화하면 위험):
- 대한과학/Agilent는 공통 수량단위 분리 엔진 자체를 스킵(하이픈 뒤가 수량이 아니라 서로 다른 실제
  제품인 회사라서).
- Thermo Fisher는 'KOLAS'로 시작하는 값만 예외로 수량단위 분리 스킵(용량범위 표기로 판단).

사전 확인(2026-08-12): 새 40개 회사 데이터에서 전역화 대상 패턴을 미리 검색한 결과 '#' 선두만
5건(DURAN 3, Saint-Gobain 2) 실제로 존재하고 나머지(견적번호/부품번호/USP/시리얼번호/모델명/끝 외/UI)는
0건이라, 이번 전역화가 새로운 오탐을 만들지 않음을 확인함.
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.52_품번_To-Be_차상위40\20_결과\260812_S-TEPS_품번_To-Be_v3.1(전역규칙화_차상위40).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_PN_CLEAN = 16
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "RXN", "TAB", "CAP", "EA", "KT", "VL", "ML", "UL", "G", "L", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)

COMPOUND_UNITS = {"AMP-EA", "KG-K", "G-F", "SET-F", "G-K"}

_NUM = r"(?:\d+(?:\.\d+)?|\.\d+)"
_QTY = rf"(?:{_NUM}[xX]{_NUM}|{_NUM})"
_UNIT = r"[A-Za-z%]+(?:-[A-Za-z%]+)?"
_SUFFIX_RE = re.compile(rf"^(?P<core>.+?)-\s*(?P<qty>{_QTY})\s*(?P<unit>{_UNIT})\Z")

_LEADING_INCH_RE = re.compile(rf"^(?P<qty>{_NUM})\s*\\?\"\s*")

# ── 전역화(회사명 무관, 내용만 맞으면 적용) ──
GLOBAL_LABEL_RULES = [
    (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^USP\s*"), "prefix"),
    (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
    (re.compile(r"\s*외\s*\Z"), "suffix"),
]
GLOBAL_UI_RE = re.compile(r"UI\Z", re.IGNORECASE)

# ── 회사별 안전장치(그대로 유지, 전역화 안 함) ──
SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
SUFFIX_ENGINE_EXCLUDE_MFRS = {"대한과학", "Agilent"}

NEW40_MFRS = {
    "Charles River", "Km", "R&D Systems", "Combi-Blocks", "Abcam", "PerkinElmer", "대정화금",
    "BioLegend", "SPL", "Saint-Gobain", "Supelco", "싸토리우스코리아", "영인에스티", "TCI", "비티알",
    "DURAN", "Promega", "DAE", "EP", "Cobetter", "MedChemExpress", "ProteinSimple", "GEMÜ",
    "Lonza", "Metrohm", "Avantor", "Beckman Coulter", "Agilent Technologies", "(주)BB", "TRC",
    "Repligen", "Millipore", "Isolab", "Siemens", "Daejung", "Gilson", "DUKSAN", "Rhawn",
    "시너지이노베이션", "Advanced Instruments",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _unit_matches(unit_up: str):
    if unit_up in COMPOUND_UNITS:
        return unit_up
    if "-" in unit_up:
        return None
    if unit_up in UNIT_WHITELIST:
        return unit_up
    for wl in UNIT_WHITELIST:
        if unit_up.startswith(wl):
            return wl
    return None


def compute_tobe(raw_pn: str, mfr: str):
    v = str(raw_pn).strip()
    removed_parts = []

    for pattern, kind in GLOBAL_LABEL_RULES:
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()
            break  # 값 하나당 라벨 규칙은 최대 1개만 적용

    if v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    m = GLOBAL_UI_RE.search(v)
    if m:
        removed_parts.append(v[m.start():])
        v = v[: m.start()]

    m = _LEADING_INCH_RE.match(v)
    if m:
        removed_parts.append(m.group(0))
        v = v[m.end():].strip()

    suffix_split_done = False
    if mfr not in SUFFIX_ENGINE_EXCLUDE_MFRS:
        if not (mfr == "Thermo Fisher Scientific" and v.upper().startswith("KOLAS")):
            m = _SUFFIX_RE.match(v)
            if m:
                unit_up = m.group("unit").upper()
                matched = _unit_matches(unit_up)
                if matched:
                    suffix_text = v[len(m.group("core")):]
                    removed_parts.append(suffix_text)
                    v = m.group("core").strip()
                    suffix_split_done = True

    memo = ""
    if mfr in SIGMA_MFRS and not suffix_split_done:
        memo = "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"

    separated_text = "".join(removed_parts)
    return v, memo, separated_text


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in NEW40_MFRS:
            continue
        tobe, memo, sep = compute_tobe(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시: {n_memo}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
