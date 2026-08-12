# -*- coding: utf-8 -*-
r"""
v3.2: v3.1에 2개 패턴만 추가.

1) 단위 화이트리스트에 M, MM 추가.
   예: 비티알 'AH-APSH-P-1000-15-15-250MM' -> 'AH-APSH-P-1000-15-15'
       비티알 'AH-APSH-P-0375-05-05-2M' -> 'AH-APSH-P-0375-05-05'
   (R&D Systems '3136-RL-01M/CF'처럼 단위 뒤에 '/CF'가 더 붙은 경우는 끝까지 안 맞아서 그대로 유지됨)
2) 전역 라벨 규칙에 '일련번호' 추가(기존 '시리얼 번호:'와 별개 표현).
   예: GEMÜ '일련번호 J20240419-Q6' -> 'J20240419-Q6'

사전 확인(2026-08-12): 상위20(1.41.51) 데이터에는 이 패턴이 0건이라 이번 수정은 차상위40(v3.x)에만
반영함.
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.52_품번_To-Be_차상위40\20_결과\260812_S-TEPS_품번_To-Be_v3.2(M_MM_일련번호_차상위40).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_PN_CLEAN = 16
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "RXN", "TAB", "CAP", "EA", "KT", "VL", "MM", "ML", "UL", "G", "L", "M", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)

COMPOUND_UNITS = {"AMP-EA", "KG-K", "G-F", "SET-F", "G-K"}

_NUM = r"(?:\d+(?:\.\d+)?|\.\d+)"
_QTY = rf"(?:{_NUM}[xX]{_NUM}|{_NUM})"
_UNIT = r"[A-Za-z%]+(?:-[A-Za-z%]+)?"
_SUFFIX_RE = re.compile(rf"^(?P<core>.+?)-\s*(?P<qty>{_QTY})\s*(?P<unit>{_UNIT})\Z")

_LEADING_INCH_RE = re.compile(rf"^(?P<qty>{_NUM})\s*\\?\"\s*")

GLOBAL_LABEL_RULES = [
    (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^USP\s*"), "prefix"),
    (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^일련\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
    (re.compile(r"\s*외\s*\Z"), "suffix"),
]
GLOBAL_UI_RE = re.compile(r"UI\Z", re.IGNORECASE)

SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
SUFFIX_ENGINE_EXCLUDE_MFRS = {"대한과학", "Agilent"}

NEW40_MFRS = {
    "Charles River", "Km", "R&D Systems", "Combi-Blocks", "Abcam", "PerkinElmer", "대정화금",
    "BioLegend", "SPL", "Saint-Gobain", "Supelco", "싸토리우스코리아", "영인에스티", "TCI", "비티알",
    "DURAN", "Promega", "DAE", "EP", "Cobetter", "MedChemExpress", "ProteinSimple", "GEMÜ",
    "Lonza", "Metrohm", "Avantor", "Beckman Coulter", "Agilent Technologies", "(주)BB", "TRC",
    "Repligen", "Millipore", "Isolab", "Siemens", "Daejung", "Gilson", "DUKSAN", "Rhawn",
    "시너지이노베이션", "Advanced Instruments",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _unit_matches(unit_up: str):
    if unit_up in COMPOUND_UNITS:
        return unit_up
    if "-" in unit_up:
        return None
    if unit_up in UNIT_WHITELIST:
        return unit_up
    for wl in UNIT_WHITELIST:
        if unit_up.startswith(wl):
            return wl
    return None


def compute_tobe(raw_pn: str, mfr: str):
    v = str(raw_pn).strip()
    removed_parts = []

    for pattern, kind in GLOBAL_LABEL_RULES:
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()
            break

    if v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    m = GLOBAL_UI_RE.search(v)
    if m:
        removed_parts.append(v[m.start():])
        v = v[: m.start()]

    m = _LEADING_INCH_RE.match(v)
    if m:
        removed_parts.append(m.group(0))
        v = v[m.end():].strip()

    suffix_split_done = False
    if mfr not in SUFFIX_ENGINE_EXCLUDE_MFRS:
        if not (mfr == "Thermo Fisher Scientific" and v.upper().startswith("KOLAS")):
            m = _SUFFIX_RE.match(v)
            if m:
                unit_up = m.group("unit").upper()
                matched = _unit_matches(unit_up)
                if matched:
                    suffix_text = v[len(m.group("core")):]
                    removed_parts.append(suffix_text)
                    v = m.group("core").strip()
                    suffix_split_done = True

    memo = ""
    if mfr in SIGMA_MFRS and not suffix_split_done:
        memo = "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"

    separated_text = "".join(removed_parts)
    return v, memo, separated_text


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in NEW40_MFRS:
            continue
        tobe, memo, sep = compute_tobe(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시: {n_memo}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
