# -*- coding: utf-8 -*-
r"""
v1.3 -> v1.4 (2차분리)
'품명 추출중'(최종 확정 품명 컬럼)과 O열(품번) 값이 완전히 동일한 26건 중,
실제로는 화학물질명/서비스명/설명일 뿐 품번이 아닌 11건을 [비품번]으로 재분류.
나머지 15건(코드가 이름칸에도 중복 기재된 것)은 [품번] 그대로 유지.
"""
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리(1차)_단위분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.3.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.42_품번_2차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번2차판정_v1.4.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_O_품번 = 15
COL_O열_판정 = 16
LABEL_ITEM = "[품번]"
LABEL_DESC = "[비품번]"

# 품명=품번 완전일치 26건 중, 실제로 코드가 아니라 이름/설명인 것으로 확인된 값
NAME_NOT_ITEM_NO = {
    "아이콘액.", "믹서기컵", "믹서기칼날", "광동 벤포파워제트액",
    "L-Alanyl-L-Glutamine", "Polysorbate 80 (HX2)", "Airsampler",
    "REPAIR", "BSC Validation", "KOLAS Certification",
}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n_reclassified = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_O열_판정).value != LABEL_ITEM:
            continue
        v = ws.cell(row=r, column=COL_O_품번).value
        s = "" if v is None else str(v).strip()
        if s in NAME_NOT_ITEM_NO:
            ws.cell(row=r, column=COL_O열_판정, value=LABEL_DESC)
            n_reclassified += 1

    wb.save(DST)
    print(f"[요약] 품명=품번 중 비품번으로 재분류: {n_reclassified}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
