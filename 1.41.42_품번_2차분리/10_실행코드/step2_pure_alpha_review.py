# -*- coding: utf-8 -*-
r"""
v1.4 -> v1.5
숫자가 전혀 없는(순수 문자) 값 187개 전수 검토 결과, 명백히 코드가 아닌
일반명사/서비스명/한국어 설명 19건을 [비품번]으로 재분류.
애매한 값(예: AALBORG, 코아텍, Northern lights 등)은 확신 없어 그대로 [품번] 유지.
"""
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.42_품번_2차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번2차판정_v1.4.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.42_품번_2차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번2차판정_v1.5.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_O_품번 = 15
COL_O열_판정 = 16
LABEL_ITEM = "[품번]"
LABEL_DESC = "[비품번]"

CONFIRMED_NOT_ITEM_NO = {
    "MISC", "SERVICE", "table", "Battery", "Calibration", "Drawer",
    "Validation", "Validation Test", "for business", "LOCAL",
    "COMPUTER", "Lab Marker",
    "KOLAS검교정", "잉크 희석제", "주문제작(품번없음)", "주문제작건",
    "청관제", "합성 표준품",
}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_O열_판정).value != LABEL_ITEM:
            continue
        v = ws.cell(row=r, column=COL_O_품번).value
        s = "" if v is None else str(v).strip()
        if s in CONFIRMED_NOT_ITEM_NO:
            ws.cell(row=r, column=COL_O열_판정, value=LABEL_DESC)
            n += 1

    wb.save(DST)
    print(f"[요약] 순수문자 검토로 비품번 재분류: {n}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
