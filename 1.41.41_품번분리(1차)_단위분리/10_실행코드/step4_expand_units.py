# -*- coding: utf-8 -*-
r"""
v1.1 -> v1.2
O열 전체 재검증 결과 확인된 EA(개당)/UL(마이크로리터)/AMP(앰플)을 승인 단위 목록에 추가하고,
아직 정제방식이 없는(=미처리) [품번] 행에 한해 단위분리를 재시도.
"""
import re
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.1.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.2.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_O_품번 = 15
COL_O열_판정 = 16
COL_O열_품번_코어 = 17
COL_O열_품번_단위 = 18
COL_O열_정제방식 = 19
LABEL_ITEM = "[품번]"

UNIT_RE = re.compile(r"^(.+)-(\d+)\s*(ML|MG|UG|KG|RXN|G|L|EA|UL|AMP)$", re.IGNORECASE)
METHOD_UNIT_SPLIT = "단위분리"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n_new = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_O열_판정).value != LABEL_ITEM:
            continue
        if ws.cell(row=r, column=COL_O열_정제방식).value:
            continue  # 이미 처리된 행은 건드리지 않음

        v = ws.cell(row=r, column=COL_O_품번).value
        s = "" if v is None else str(v).strip()
        m = UNIT_RE.match(s)
        if not m:
            continue

        core = m.group(1)
        unit = (m.group(2) + m.group(3)).upper()
        ws.cell(row=r, column=COL_O열_품번_코어, value=core)
        ws.cell(row=r, column=COL_O열_품번_단위, value=unit)
        ws.cell(row=r, column=COL_O열_정제방식, value=METHOD_UNIT_SPLIT)
        n_new += 1

    wb.save(DST)
    print(f"[요약] EA/UL/AMP 추가로 새로 단위분리된 행: {n_new}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
