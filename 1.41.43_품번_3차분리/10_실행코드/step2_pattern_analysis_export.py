# -*- coding: utf-8 -*-
r"""
남은 13,678건 품번(잠정)에 대한 패턴 분석 결과를 엑셀로 export.
- 제조사 분포 전체
- 형태패턴(문자=A, 숫자=9 치환) 분포 전체
- Merck 점구분 형식 교차검증 상세
"""
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43_품번_3차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번3차판정_v1.6.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43_품번_3차분리\20_결과\260810_품번_패턴분석_결과.xlsx")

SHEET = "Steps_중복제거_32359"

HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)


def shape(s: str) -> str:
    out = []
    for ch in s:
        if ch.isalpha():
            out.append("A")
        elif ch.isdigit():
            out.append("9")
        else:
            out.append(ch)
    return "".join(out)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb_src = openpyxl.load_workbook(SRC, data_only=False, read_only=True)
    ws_src = wb_src[SHEET]

    values = []
    manufacturers = []
    for row in ws_src.iter_rows(min_row=5):
        o = row[14].value
        s1, s2, s3 = row[15].value, row[16].value, row[17].value
        if s1 != "[품번]" or s2 or s3 or o is None:
            continue
        s = str(o).strip()
        values.append(s)
        manufacturers.append((row[21].value or "").strip())

    total = len(values)

    mfr_counter = Counter(m.lower() if m else "(없음)" for m in manufacturers)
    shape_counter = Counter(shape(v) for v in values)

    dot_shapes = {"9.99999.9999", "9999.9999", "9.9999.999"}
    dot_examples = []
    dot_mfr_counter = Counter()
    for v, m in zip(values, manufacturers):
        if shape(v) in dot_shapes:
            dot_mfr_counter[m.lower() if m else "(없음)"] += 1
            if len(dot_examples) < 30:
                dot_examples.append((v, m))

    out = openpyxl.Workbook()

    # ---- 시트1: 설명 ----
    ws = out.active
    ws.title = "설명"
    ws["A1"] = "O열(품번) 패턴 분석 결과"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        f"분석 대상: 1~3차 판정에서 값없음/CAS의심/비품번/단위분리/기호제거 어디에도 해당하지 않고 남은 '품번(잠정)' {total:,}건",
        "",
        "1. 형태패턴 분석",
        "  각 값의 문자를 A, 숫자를 9로 치환해서 '형태 지문'을 만들고 빈도를 집계함.",
        f"  전체 고유 형태 지문: {len(shape_counter):,}개",
        "  '형태패턴_분포' 시트에 전체 목록(빈도순) 수록.",
        "",
        "2. 제조사 분포 분석",
        "  같은 행의 '제조사' 열 값을 대소문자 무시하고 집계.",
        f"  제조사 값이 있는 행: {total - mfr_counter.get('(없음)', 0):,}건 / 없는 행: {mfr_counter.get('(없음)', 0):,}건",
        "  '제조사_분포' 시트에 전체 목록(빈도순) 수록.",
        "",
        "3. 형태-제조사 교차검증 사례 (Merck 점구분 형식)",
        "  형태 지문이 '9.99999.9999' / '9999.9999' / '9.9999.999' (점으로 구분된 숫자코드)인 행을 뽑아",
        "  제조사 열과 대조한 결과, Merck 계열(Merck/MERCK/merck/Merck Millipore)이 큰 비중을 차지함을 확인.",
        "  즉 특정 형태 패턴이 특정 제조사의 고유 카탈로그 번호 체계와 실제로 연결되어 있음.",
        "  'Merck_점구분_교차검증' 시트에 제조사별 집계 및 값 예시 수록.",
        "",
        "다음 단계 후보: 제조사별로 그룹핑해서 각 제조사 안에서 형태가 크게 벗어나는 이상값을 찾는 방식",
        "(현재는 분석/보고 목적으로만 작성 — 원본 파일은 변경하지 않음)",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110

    # ---- 시트2: 제조사_분포 ----
    ws2 = out.create_sheet("제조사_분포")
    ws2.append(["제조사(소문자 정규화)", "건수", "비율(%)", "누적비율(%)"])
    style_header(ws2, 1, 4)
    cum = 0
    for mfr, c in mfr_counter.most_common():
        cum += c
        ws2.append([mfr, c, round(c / total * 100, 2), round(cum / total * 100, 2)])
    autosize(ws2, [30, 10, 12, 14])
    ws2.freeze_panes = "A2"

    # ---- 시트3: 형태패턴_분포 ----
    ws3 = out.create_sheet("형태패턴_분포")
    ws3.append(["형태 지문(문자=A, 숫자=9)", "건수", "비율(%)", "누적비율(%)"])
    style_header(ws3, 1, 4)
    cum = 0
    for sh, c in shape_counter.most_common():
        cum += c
        ws3.append([sh, c, round(c / total * 100, 2), round(cum / total * 100, 2)])
    autosize(ws3, [30, 10, 12, 14])
    ws3.freeze_panes = "A2"

    # ---- 시트4: Merck_점구분_교차검증 ----
    ws4 = out.create_sheet("Merck_점구분_교차검증")
    ws4["A1"] = "형태 지문이 점구분 숫자코드(9.99999.9999 등)인 행의 제조사 분포"
    ws4["A1"].font = HEADER_FONT
    ws4.append(["제조사(소문자 정규화)", "건수"])
    style_header(ws4, 3, 2)
    r = 4
    for mfr, c in dot_mfr_counter.most_common():
        ws4.cell(row=r, column=1, value=mfr)
        ws4.cell(row=r, column=2, value=c)
        r += 1
    r += 2
    ws4.cell(row=r, column=1, value="값 예시 (품번 / 제조사)").font = HEADER_FONT
    r += 1
    ws4.cell(row=r, column=1, value="품번")
    ws4.cell(row=r, column=2, value="제조사")
    style_header(ws4, r, 2)
    r += 1
    for v, m in dot_examples:
        ws4.cell(row=r, column=1, value=v)
        ws4.cell(row=r, column=2, value=m)
        r += 1
    autosize(ws4, [30, 20])

    DST.parent.mkdir(parents=True, exist_ok=True)
    out.save(DST)
    print(f"[저장] {DST}")
    print(f"[요약] 대상 {total}건 / 제조사 고유값 {len(mfr_counter)}개 / 형태패턴 고유값 {len(shape_counter)}개")


if __name__ == "__main__":
    main()
