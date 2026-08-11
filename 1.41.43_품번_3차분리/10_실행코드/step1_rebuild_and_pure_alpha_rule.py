# -*- coding: utf-8 -*-
r"""
v1.3(1차 결과, 오염 없는 상태) -> v1.6

1) O열_판정 -> O열_1차판정으로 이름 정리 (원래 1차 판정 그대로, 손대지 않음)
2) O열_2차판정 신규 열: 2차분리에서 실제로 재분류했던 33건("품명=품번" 12건 +
   순수문자 확인 21건)을 여기로 옮김 (기존처럼 1차판정 열을 직접 덮어쓰지 않음)
3) O열_3차판정 신규 열: "품번(품목 번호)인데 숫자가 하나도 없으면 애초에 번호가
   아니다"라는 원칙을 전수 적용 -> 1차판정이 [품번]이고 2차에서 안 건드린 행 중
   숫자가 전혀 없는 값 전부 [비품번]으로 분류 (예외 없이 전수 적용)
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.41_품번분리(1차)_단위분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번1차판정_v1.3.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43_품번_3차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번3차판정_v1.6.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_O_품번 = 15
COL_O열_1차판정 = 16   # 기존 O열_판정, 이름만 바꿈
INSERT_AT = 17          # 1차판정 다음
N_NEW_COLS = 2
LABEL_ITEM = "[품번]"
LABEL_DESC = "[비품번]"

# 2차분리에서 실제 재분류했던 값 전체(두 차례 검토분 합침)
STAGE2_RECLASSIFIED = {
    # 품명=품번 완전일치 12건
    "아이콘액.", "믹서기컵", "믹서기칼날", "광동 벤포파워제트액",
    "L-Alanyl-L-Glutamine", "Polysorbate 80 (HX2)", "Airsampler",
    "REPAIR", "BSC Validation", "KOLAS Certification",
    # 순수문자 전수검토 확정 21건
    "MISC", "SERVICE", "table", "Battery", "Calibration", "Drawer",
    "Validation", "Validation Test", "for business", "LOCAL",
    "COMPUTER", "Lab Marker",
    "KOLAS검교정", "잉크 희석제", "주문제작(품번없음)", "주문제작건",
    "청관제", "합성 표준품",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    ws.cell(row=4, column=COL_O열_1차판정, value="O열_1차판정")

    computed = {}  # row -> (stage2_label, stage3_label)
    n_stage2 = n_stage3 = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_O열_1차판정).value != LABEL_ITEM:
            continue
        v = ws.cell(row=r, column=COL_O_품번).value
        s = "" if v is None else str(v).strip()

        if s in STAGE2_RECLASSIFIED:
            computed[r] = (LABEL_DESC, "")
            n_stage2 += 1
            continue

        if s and not any(ch.isdigit() for ch in s):
            computed[r] = ("", LABEL_DESC)
            n_stage3 += 1

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    for i in range(N_NEW_COLS):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].hidden = False
        ws.column_dimensions[letter].width = 14

    ws.cell(row=4, column=INSERT_AT, value="O열_2차판정")
    ws.cell(row=4, column=INSERT_AT + 1, value="O열_3차판정")

    for r, (s2, s3) in computed.items():
        if s2:
            ws.cell(row=r, column=INSERT_AT, value=s2)
        if s3:
            ws.cell(row=r, column=INSERT_AT + 1, value=s3)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] O열_2차판정(재분류): {n_stage2}건")
    print(f"[요약] O열_3차판정(순수문자 전수 적용): {n_stage3}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
