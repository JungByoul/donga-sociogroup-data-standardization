# -*- coding: utf-8 -*-
"""[검증 전용] 육안 확인용 무작위 샘플 출력."""
import random
import openpyxl

p = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260811_S-TEPS_품번_To-Be_v1.0.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
ws = wb["Steps_중복제거_32359"]

COL_ITEM = 13
COL_PN = 15
COL_TOBE = 17
COL_MEMO = 18
COL_MFR = 21

rows_with_tobe = [r for r in range(5, ws.max_row + 1) if ws.cell(row=r, column=COL_TOBE).value]
random.seed(42)
sample = random.sample(rows_with_tobe, 15)
sample.sort()

print(f"{'품목명':40s} | {'제조사':20s} | {'원본품번':20s} | {'To-Be':15s} | 비고")
for r in sample:
    item = str(ws.cell(row=r, column=COL_ITEM).value)[:38]
    mfr = str(ws.cell(row=r, column=COL_MFR).value)
    pn = str(ws.cell(row=r, column=COL_PN).value)
    tobe = str(ws.cell(row=r, column=COL_TOBE).value)
    memo = ws.cell(row=r, column=COL_MEMO).value or ""
    print(f"{item:40s} | {mfr:20s} | {pn:20s} | {tobe:15s} | {memo}")
