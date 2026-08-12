# -*- coding: utf-8 -*-
r"""
v2.2: v2.1 로직에 아래 보완 사항 추가 (전부 사용자 확인 완료, 여전히 TOP20 범위 내).

1) 화이트리스트에 RXN/KT/TAB/VL 추가 (기존 KG/MG/UG/NG/AMP/PAK/EA/ML/UL/G/L/%).
2) 수량 표기 확장:
   - 소수점 앞에 0(정수부)이 없는 경우도 인식 (예: '-.5ML').
   - 'NxM단위' 곱셈 표기 인식 (예: '-5x10ML', '-4X100ML', '-10X2ML').
3) 하이픈을 포함한 복합단위 인정: 'AMP-EA', 'KG-K' (Sigma/Sigma-Aldrich, 각 1건/2건).
   예: '45-T6508-10AMP-EA' -> '45-T6508' / 분리텍스트 '-10AMP-EA'
       'W278430-1KG-K' -> 'W278430' / 분리텍스트 '-1KG-K'
4) Agilent 'UI' 접미사 전용 처리 — Agilent는 하이픈 뒤 숫자가 수량이 아니라 서로 다른 실제
   제품임이 1단계에서 이미 확인된 회사라, 공통 수량단위 분리 엔진을 그대로 적용하면 안 됨.
   'UI' 글자만 장식 기호로 보고 떼어내고 숫자는 그대로 유지.
   예: '123-0364UI' -> '123-0364' / 분리텍스트 'UI'  (하이픈+숫자는 안 건드림)

v2.1과 동일하게 원본(1.41.50_v0.5)에서 새로 빌드. 컬럼 구성(품번_To-Be/비고/분리텍스트)도 동일.
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260812_S-TEPS_품번_To-Be_v2.2(복합단위_곱셈표기_Agilent UI_단위확장).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_PN_CLEAN = 16
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "RXN", "TAB", "EA", "KT", "VL", "ML", "UL", "G", "L", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)

COMPOUND_UNITS = {"AMP-EA", "KG-K"}  # 단위 자체가 하이픈을 포함하는 예외 케이스 (정확히 일치할 때만 인정)

# qty: 일반 정수/소수(1, 2.5) | 앞자리 없는 소수(.5) | NxM 곱셈(5x10, 4X100)
_QTY = r"(?:\d+[xX]\d+(?:\.\d+)?|\d+(?:\.\d+)?|\.\d+)"
_UNIT = r"[A-Za-z%]+(?:-[A-Za-z%]+)?"  # 뒤에 '-단어' 하나까지 복합단위로 허용(AMP-EA, KG-K 검증은 별도)
_SUFFIX_RE = re.compile(rf"^(?P<core>.+?)-(?P<qty>{_QTY})\s*(?P<unit>{_UNIT})\Z")

SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
HASH_STRIP_MFRS = {"대한과학", "Cell Signaling Technology", "Sigma", "Sigma-Aldrich"}
BRACKET_STRIP_MFRS = {"Sartorius"}
SUFFIX_ENGINE_EXCLUDE_MFRS = {"대한과학", "Agilent"}  # Agilent는 별도 UI 전용 규칙으로만 처리

LABEL_RULES = {
    "Sartorius": (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    "Agilent": (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    "USP": (re.compile(r"^USP\s*"), "prefix"),
    "Mettler Toledo": (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    "Cytiva": (re.compile(r"\s*외\s*\Z"), "suffix"),
    "Eppendorf": (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
}

AGILENT_UI_RE = re.compile(r"UI\Z", re.IGNORECASE)

TOP20_MFRS = {
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _unit_matches(unit_up: str):
    """단위 문자열이 화이트리스트에 맞는지 판정. (정확히 일치 / 화이트리스트로 시작 / 복합단위 정확일치)"""
    if unit_up in COMPOUND_UNITS:
        return unit_up
    if "-" in unit_up:
        return None  # 하이픈 포함 단위는 COMPOUND_UNITS에 정확히 등록된 것만 인정(그 외는 오탐 방지)
    if unit_up in UNIT_WHITELIST:
        return unit_up
    for wl in UNIT_WHITELIST:
        if unit_up.startswith(wl):
            return wl
    return None


def compute_tobe(raw_pn: str, mfr: str):
    v = str(raw_pn).strip()
    removed_parts = []

    rule = LABEL_RULES.get(mfr)
    if rule:
        pattern, kind = rule
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()

    if mfr in HASH_STRIP_MFRS and v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if mfr in BRACKET_STRIP_MFRS and v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    if mfr == "Agilent":
        m = AGILENT_UI_RE.search(v)
        if m:
            removed_parts.append(v[m.start():])
            v = v[: m.start()]

    suffix_split_done = False
    if mfr not in SUFFIX_ENGINE_EXCLUDE_MFRS:
        if not (mfr == "Thermo Fisher Scientific" and v.upper().startswith("KOLAS")):
            m = _SUFFIX_RE.match(v)
            if m:
                unit_up = m.group("unit").upper()
                matched = _unit_matches(unit_up)
                if matched:
                    suffix_text = v[len(m.group("core")):]
                    removed_parts.append(suffix_text)
                    v = m.group("core").strip()
                    suffix_split_done = True

    memo = ""
    if mfr in SIGMA_MFRS and not suffix_split_done:
        memo = "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"

    separated_text = "".join(removed_parts)
    return v, memo, separated_text


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in TOP20_MFRS:
            continue
        tobe, memo, sep = compute_tobe(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시(Sigma 계열 미분리): {n_memo}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
