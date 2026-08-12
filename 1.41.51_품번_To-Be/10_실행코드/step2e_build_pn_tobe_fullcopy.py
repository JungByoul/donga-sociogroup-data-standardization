# -*- coding: utf-8 -*-
r"""
1.41.50_v0.5 원본 파일을 통째로 복사(포맷/다른 시트/스타일 전부 유지)한 뒤,
Steps_중복제거_32359 시트의 P열(품번_정리) 오른쪽에 '품번_To-Be' / '품번_To-Be_비고' 2개 열을
실제로 삽입한다(값 전용으로 새로 만드는 대신, 원본 파일 자체를 편집).

사전 점검 결과(step2a/step2a2):
- 이 시트 안 수식은 M2 하나뿐: =COUNTIF(AI6:AI32354,"*!*")  -> AI(35열)가 삽입지점(17) 이후라 AK(37열)로 보정 필요
- 다른 시트가 이 시트를 참조하는 수식 없음
- 정의된 이름 'bw'는 LAMBDA 함수 정의라 셀 참조와 무관, 손댈 필요 없음
- 병합된 셀 없음
따라서 컬럼 삽입 후 M2 수식 하나만 보정하면 안전함.
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260811_S-TEPS_품번_To-Be_v2.0(원본유지).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15            # O: 품번
COL_PN_CLEAN = 16      # P: 품번_정리
INSERT_AT = 17         # Q부터 밀림
N_NEW_COLS = 2
COL_MFR_CLEAN_OLD = 20  # 삽입 전 '제조사\n정리' 위치(원본 기준, pandas 읽기용이라 변하지 않음)

UNIT_WHITELIST = {"G", "KG", "MG", "UG", "NG", "ML", "L", "UL", "%"}
_SUFFIX_RE = re.compile(r"^(?P<core>.+?)-(?P<qty>\d+(\.\d+)?)\s*(?P<unit>[A-Za-z%]*)$")

SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
HASH_STRIP_MFRS = {"대한과학", "Cell Signaling Technology"}
BRACKET_STRIP_MFRS = {"Sartorius"}

TOP20_MFRS = {
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def compute_tobe(raw_pn: str, mfr: str):
    v = str(raw_pn).strip()
    if mfr in SIGMA_MFRS:
        m = _SUFFIX_RE.match(v)
        if m and m.group("unit").upper() in UNIT_WHITELIST:
            return m.group("core").strip(), ""
        return v, "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"
    if mfr in HASH_STRIP_MFRS:
        if v.startswith("#"):
            return v[1:].strip(), ""
        return v, ""
    if mfr in BRACKET_STRIP_MFRS:
        if v.startswith("[") and v.endswith("]"):
            return v[1:-1].strip(), ""
        return v, ""
    return v, ""


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    # 1) 컬럼 삽입 전, 보정 대상 수식 수집
    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    # 2) 컬럼 삽입 (P열 다음 = Q열 위치에 2칸)
    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    # 3) 수식 보정 (openpyxl insert_cols는 셀은 밀어주지만 수식 텍스트 안의 참조는 안 고쳐줌)
    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    # 4) 새 컬럼 헤더/서식
    for i in range(N_NEW_COLS):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = 22 if i == 0 else 32
    ws.cell(row=HEADER_ROW, column=INSERT_AT, value="품번_To-Be")
    ws.cell(row=HEADER_ROW, column=INSERT_AT + 1, value="품번_To-Be_비고")

    # 5) 규칙 적용 — 원본(SRC)에서 pandas로 읽어 대상 행/값 판별 (행 번호는 열 삽입과 무관하게 그대로)
    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in TOP20_MFRS:
            continue
        tobe, memo = compute_tobe(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시: {n_memo}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
