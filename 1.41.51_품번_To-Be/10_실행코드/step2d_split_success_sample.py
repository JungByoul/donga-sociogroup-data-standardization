# -*- coding: utf-8 -*-
"""[검증 전용] Sigma/Sigma-Aldrich 핵심코드 분리 성공 사례 샘플."""
import openpyxl

p = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260811_S-TEPS_품번_To-Be_v1.0.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
ws = wb["Steps_중복제거_32359"]

COL_ITEM = 13
COL_PN = 15
COL_TOBE = 17
COL_MEMO = 18
COL_MFR = 21

print(f"{'품목명':35s} | {'제조사':15s} | {'원본품번':18s} | To-Be")
n = 0
for r in range(5, ws.max_row + 1):
    mfr = ws.cell(row=r, column=COL_MFR).value
    if mfr not in ("Sigma", "Sigma-Aldrich"):
        continue
    tobe = ws.cell(row=r, column=COL_TOBE).value
    memo = ws.cell(row=r, column=COL_MEMO).value
    pn = ws.cell(row=r, column=COL_PN).value
    if tobe and not memo and str(tobe) != str(pn).strip():
        item = str(ws.cell(row=r, column=COL_ITEM).value)[:33]
        print(f"{item:35s} | {mfr:15s} | {str(pn):18s} | {tobe}")
        n += 1
        if n >= 12:
            break
