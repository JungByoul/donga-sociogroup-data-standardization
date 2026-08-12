# -*- coding: utf-8 -*-
r"""
v2.1: v2.0(품번_To-Be / 품번_To-Be_비고) 로직을 아래처럼 보완.

사용자 검토(v2.0) 피드백 반영 사항:
1) 수량단위 접미사 분리(기존 Sigma/Sigma-Aldrich 전용)를 TOP20 전 회사 공통 엔진으로 확장.
   화이트리스트에 EA/AMP/PAK 추가(기존 G/KG/MG/UG/NG/ML/L/UL/% 유지).
   단위가 화이트리스트와 '정확히 일치'하지 않아도 화이트리스트 단어로 '시작'하면 인정
   (예: Millipore '100GMCN' -> G로 인식, Sigma-Aldrich '50UGCN' -> UG로 인식).
2) 회사별 라벨 프리픽스/서픽스 제거 추가:
   - Sartorius: '견적번호', '견적서 번호 : ' 등 견적 라벨 제거
   - Agilent: '부품 번호:' 라벨 제거
   - USP: 'USP ' 라벨 제거
   - Mettler Toledo: '시리얼 번호:' 라벨 제거
   - Cytiva: 끝의 ' 외' 제거
   - Eppendorf: '모델명: ' 라벨 제거
   (라벨 없는 다른 회사는 이 규칙 미적용 - 회사 코드에 명시된 라벨만 대상)
3) 예외 처리(사용자 확인 완료):
   - Thermo Fisher 'KOLAS...' 로 시작하는 값은 용량 범위 표기로 판단, 공통단위분리 미적용
   - 대한과학은 원래 '#' 제거만 하던 회사라 공통단위분리 자체를 적용하지 않음(# 제거만 유지)
4) 신규 컬럼 '품번_To-Be_분리텍스트' 추가(S열) - 원본에서 잘려나간 부분을 그대로 기록
   (예: '3232-1EA' -> To-Be '3232', 분리텍스트 '-1EA')

사전 점검 결과(step2a/step2a2, v2.0과 동일 파일 대상이라 재사용):
- 시트 내 수식은 M2 하나: =COUNTIF(AI6:AI32354,"*!*") -> 열 삽입 개수만큼 보정 필요(2->3으로 변경)
- 다른 시트가 이 시트를 참조하는 수식 없음 / 정의된 이름 'bw'는 셀 참조 무관 / 병합 셀 없음
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260812_S-TEPS_품번_To-Be_v2.1(공통단위분리_라벨제거_분리텍스트열추가).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15             # O: 품번
COL_PN_CLEAN = 16       # P: 품번_정리
INSERT_AT = 17          # Q부터 밀림
N_NEW_COLS = 3           # Q=품번_To-Be, R=품번_To-Be_비고, S=품번_To-Be_분리텍스트
COL_MFR_CLEAN_OLD = 20   # 삽입 전 '제조사\n정리' 위치(원본 기준, pandas 읽기용이라 변하지 않음)

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "EA", "ML", "UL", "G", "L", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)  # 긴 토큰부터 매칭해야 'KG'가 'G'로 잘못 인식되지 않음

_SUFFIX_RE = re.compile(r"^(?P<core>.+?)-(?P<qty>\d+(\.\d+)?)\s*(?P<unit>[A-Za-z%]*)\Z")

SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
HASH_STRIP_MFRS = {"대한과학", "Cell Signaling Technology", "Sigma", "Sigma-Aldrich"}
BRACKET_STRIP_MFRS = {"Sartorius"}
SUFFIX_ENGINE_EXCLUDE_MFRS = {"대한과학"}  # 공통단위분리 엔진 자체를 적용하지 않는 회사

LABEL_RULES = {
    "Sartorius": (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    "Agilent": (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    "USP": (re.compile(r"^USP\s*"), "prefix"),
    "Mettler Toledo": (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    "Cytiva": (re.compile(r"\s*외\s*\Z"), "suffix"),
    "Eppendorf": (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
}

TOP20_MFRS = {
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def compute_tobe(raw_pn: str, mfr: str):
    v = str(raw_pn).strip()
    removed_parts = []

    rule = LABEL_RULES.get(mfr)
    if rule:
        pattern, kind = rule
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()

    if mfr in HASH_STRIP_MFRS and v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if mfr in BRACKET_STRIP_MFRS and v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    suffix_split_done = False
    if mfr not in SUFFIX_ENGINE_EXCLUDE_MFRS:
        if not (mfr == "Thermo Fisher Scientific" and v.upper().startswith("KOLAS")):
            m = _SUFFIX_RE.match(v)
            if m:
                unit_up = m.group("unit").upper()
                matched = unit_up if unit_up in UNIT_WHITELIST else next(
                    (w for w in UNIT_WHITELIST if unit_up.startswith(w)), None
                )
                if matched:
                    suffix_text = v[len(m.group("core")):]
                    removed_parts.append(suffix_text)
                    v = m.group("core").strip()
                    suffix_split_done = True

    memo = ""
    if mfr in SIGMA_MFRS and not suffix_split_done:
        memo = "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"

    separated_text = "".join(removed_parts)
    return v, memo, separated_text


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in TOP20_MFRS:
            continue
        tobe, memo, sep = compute_tobe(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시(Sigma 계열 미분리): {n_memo}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
