# -*- coding: utf-8 -*-
r"""
[분석 전용] 컬럼 삽입 전 수식 안전성 점검.
- Steps_중복제거_32359 시트 안의 수식 개수/샘플
- 다른 시트에서 이 시트를 참조하는 수식이 있는지
- 워크북에 정의된 이름(Defined Names)이 있는지
"""
import openpyxl

SRC = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx"
SHEET = "Steps_중복제거_32359"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)

    ws = wb[SHEET]
    n_formula = 0
    samples = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                n_formula += 1
                if len(samples) < 15:
                    samples.append((cell.coordinate, cell.value))
    print(f"[{SHEET}] 자체 수식 셀 개수: {n_formula}")
    for coord, f in samples:
        print(" ", coord, f)

    print()
    print("=== 다른 시트에서 이 시트를 참조하는 수식 확인 ===")
    for sn in wb.sheetnames:
        if sn == SHEET:
            continue
        other = wb[sn]
        n_ref = 0
        ref_samples = []
        for row in other.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and SHEET in cell.value:
                    n_ref += 1
                    if len(ref_samples) < 5:
                        ref_samples.append((cell.coordinate, cell.value))
        if n_ref:
            print(f"  [{sn}] {SHEET} 참조 수식: {n_ref}건")
            for coord, f in ref_samples:
                print("    ", coord, f)

    print()
    print("=== 워크북 정의된 이름(Defined Names) ===")
    try:
        names = list(wb.defined_names.keys())
    except AttributeError:
        names = list(wb.defined_names)
    print(names if names else "(없음)")

    print()
    print("=== 병합된 셀 범위(Merged Cells) 개수 ===")
    print(len(ws.merged_cells.ranges))


if __name__ == "__main__":
    main()
