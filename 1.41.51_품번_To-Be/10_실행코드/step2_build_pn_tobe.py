# -*- coding: utf-8 -*-
r"""
1.41.50_v0.5의 Steps_중복제거_32359 시트만 값(고정값, 수식 아님)으로 떼어내서 새 파일을 만들고,
P열(품번\n정리, 16번째 컬럼) 오른쪽에 '품번_To-Be' / '품번_To-Be_비고' 2개 컬럼을 추가한다.

시트 하나만 분리하면 다른 시트를 참조하는 수식은 어차피 깨지므로(#REF!), 아예 처음부터
수식이 아니라 현재 계산된 값으로 복사한다. 원본 1.41.50 파일은 전혀 건드리지 않는다.

규칙(1단계 분석 보고서 '260811_품번_To-Be_1단계_제조사별_패턴분석_v1.0.docx'에서 확정):
- Sigma / Sigma-Aldrich: '핵심코드-수량단위' 구조에서 단위가 도량형(G/KG/MG/UG/NG/ML/L/UL/%)이면
  핵심코드만 To-Be로 채택. 매칭 안 되면 트림된 원본 유지 + 비고.
- 대한과학 / Cell Signaling Technology: 트림 + 선두 '#' 기호 제거.
- Sartorius: 트림 + 전체가 '[...]'로 감싸진 경우 대괄호 제거.
- 그 외 상위 20개 제조사(Merck Millipore, Thermo Fisher Scientific, Agilent, 삼전순약공업,
  Corning, USP, Mettler Toledo, Invitrogen, Waters, Cytiva, 유코, Roche, Eppendorf, BD, Gibco):
  트림만(구조 변경 없음). Agilent/Corning은 접미사가 수량이 아니라 서로 다른 실제 제품임을
  원문 대조로 확인했으므로 절대 분리하지 않음.
- 상위 20개 제조사가 아니거나 품번이 없는 행은 공란 유지(이번 단계 범위 밖).
"""
import re
from pathlib import Path

import openpyxl
import pandas as pd

SRC = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx"
SHEET = "Steps_중복제거_32359"
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260811_S-TEPS_품번_To-Be_v1.0.xlsx")

HEADER_ROW = 4  # 엑셀 기준(1-indexed)
DATA_START_ROW = 5

COL_PN = 15          # O열: 품번
COL_PN_CLEAN = 16    # P열: 품번_정리
INSERT_AT = 17       # P열 다음 = Q열
COL_MFR_CLEAN_OLD = 20  # 원본 기준 '제조사\n정리' 열(1-based, 삽입 전 위치)

UNIT_WHITELIST = {"G", "KG", "MG", "UG", "NG", "ML", "L", "UL", "%"}
_SUFFIX_RE = re.compile(r"^(?P<core>.+?)-(?P<qty>\d+(\.\d+)?)\s*(?P<unit>[A-Za-z%]*)$")

SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
HASH_STRIP_MFRS = {"대한과학", "Cell Signaling Technology"}
BRACKET_STRIP_MFRS = {"Sartorius"}

TOP20_MFRS = {
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
}


def compute_tobe(raw_pn: str, mfr: str):
    """returns (tobe_value, memo)"""
    v = str(raw_pn).strip()

    if mfr in SIGMA_MFRS:
        m = _SUFFIX_RE.match(v)
        if m and m.group("unit").upper() in UNIT_WHITELIST:
            return m.group("core").strip(), ""
        return v, "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"

    if mfr in HASH_STRIP_MFRS:
        if v.startswith("#"):
            return v[1:].strip(), ""
        return v, ""

    if mfr in BRACKET_STRIP_MFRS:
        if v.startswith("[") and v.endswith("]"):
            return v[1:-1].strip(), ""
        return v, ""

    # 나머지 상위20: 트림만
    return v, ""


def main():
    wb_src = openpyxl.load_workbook(SRC, data_only=True)  # 계산된 값만 사용
    ws_src = wb_src[SHEET]

    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = SHEET

    max_row = ws_src.max_row
    max_col = ws_src.max_column

    # 1) 원본 값 복사 (P열까지 그대로, Q열부터는 한 칸 밀어서 복사)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws_src.cell(row=r, column=c).value
            dst_c = c if c <= COL_PN_CLEAN else c + 1
            ws_dst.cell(row=r, column=dst_c, value=val)

    # 2) 헤더 추가
    ws_dst.cell(row=HEADER_ROW, column=INSERT_AT, value="품번_To-Be")
    ws_dst.cell(row=HEADER_ROW, column=INSERT_AT + 1, value="품번_To-Be_비고")

    # 3) 규칙 적용 (제조사/품번은 pandas로 읽어서 대상 행 판별 후 openpyxl에 기록)
    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in TOP20_MFRS:
            continue
        tobe, memo = compute_tobe(pn, mfr)
        ws_dst.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws_dst.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1

    # 4) 열 너비 대략 복사(가독성)
    for c in range(1, max_col + 2):
        src_c = c if c <= COL_PN_CLEAN else c - 1
        try:
            src_letter = openpyxl.utils.get_column_letter(src_c)
            dst_letter = openpyxl.utils.get_column_letter(c)
            if src_letter in ws_src.column_dimensions:
                ws_dst.column_dimensions[dst_letter].width = ws_src.column_dimensions[src_letter].width
        except Exception:
            pass
    ws_dst.column_dimensions[openpyxl.utils.get_column_letter(INSERT_AT)].width = 20
    ws_dst.column_dimensions[openpyxl.utils.get_column_letter(INSERT_AT + 1)].width = 30

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb_dst.save(DST)

    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건 (상위20 제조사 + 품번보유 행 대상)")
    print(f"[요약] 비고(핵심코드 분리 실패) 표시: {n_memo}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
