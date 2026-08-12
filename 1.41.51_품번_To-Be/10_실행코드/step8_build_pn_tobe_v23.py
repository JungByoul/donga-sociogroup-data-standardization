# -*- coding: utf-8 -*-
r"""
v2.3: v2.2 로직에 아래 보완 사항 추가 (전부 사용자 확인 완료, 여전히 TOP20 범위 내).

1) 화이트리스트에 CAP 추가 (기존 KG/MG/UG/NG/AMP/PAK/RXN/TAB/EA/KT/VL/ML/UL/G/L/%).
2) 복합단위(하이픈 포함) 3종 추가: 'G-F', 'SET-F', 'G-K' (기존 AMP-EA, KG-K).
   예: '11009-100G-F' -> '11009', '87574-1SET-F' -> '87574', 'W200220-100G-K' -> 'W200220'
3) NxM 곱셈표기에서 뒷자리 숫자도 소수(.5 등)만 있는 경우 인식.
   예: '646563-10X.5ML' -> '646563'
4) 하이픈 뒤에 공백이 있는 경우도 인식. 예: 'R0278- 50ML' -> 'R0278'
5) 신규: 앞쪽 수량+인치기호(") 제거 규칙 (기존 규칙은 전부 '뒤쪽 접미사' 제거였는데, 이건 유일하게
   '앞쪽 접두사'로 수량이 붙는 케이스). Merck Millipore '0.25" Clamps and gaskets'처럼 하이픈도 없고
   핵심 코드라 부를 게 없는 서술형 값 — 맨 앞의 '숫자+"' 자체를 수량/단위로 보고 제거.
   예: '0.25" Clamps and gaskets' -> 'Clamps and gaskets' (분리텍스트 '0.25" ')
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260812_S-TEPS_품번_To-Be_v2.3(복합단위추가_CAP_공백하이픈_선두인치기호).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_PN_CLEAN = 16
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "RXN", "TAB", "CAP", "EA", "KT", "VL", "ML", "UL", "G", "L", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)

COMPOUND_UNITS = {"AMP-EA", "KG-K", "G-F", "SET-F", "G-K"}

# qty: 정수/소수 | 앞자리 없는 소수 | NxM 곱셈(뒷자리도 정수/소수/앞자리없는소수 허용)
_NUM = r"(?:\d+(?:\.\d+)?|\.\d+)"
_QTY = rf"(?:{_NUM}[xX]{_NUM}|{_NUM})"
_UNIT = r"[A-Za-z%]+(?:-[A-Za-z%]+)?"
_SUFFIX_RE = re.compile(rf"^(?P<core>.+?)-\s*(?P<qty>{_QTY})\s*(?P<unit>{_UNIT})\Z")

_LEADING_INCH_RE = re.compile(rf"^(?P<qty>{_NUM})\s*\\?\"\s*")  # 실제 데이터엔 인치기호가 \" (백슬래시+따옴표)로 저장됨

SIGMA_MFRS = {"Sigma", "Sigma-Aldrich"}
HASH_STRIP_MFRS = {"대한과학", "Cell Signaling Technology", "Sigma", "Sigma-Aldrich"}
BRACKET_STRIP_MFRS = {"Sartorius"}
SUFFIX_ENGINE_EXCLUDE_MFRS = {"대한과학", "Agilent"}

LABEL_RULES = {
    "Sartorius": (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    "Agilent": (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    "USP": (re.compile(r"^USP\s*"), "prefix"),
    "Mettler Toledo": (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    "Cytiva": (re.compile(r"\s*외\s*\Z"), "suffix"),
    "Eppendorf": (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
}

AGILENT_UI_RE = re.compile(r"UI\Z", re.IGNORECASE)

TOP20_MFRS = {
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _unit_matches(unit_up: str):
    if unit_up in COMPOUND_UNITS:
        return unit_up
    if "-" in unit_up:
        return None
    if unit_up in UNIT_WHITELIST:
        return unit_up
    for wl in UNIT_WHITELIST:
        if unit_up.startswith(wl):
            return wl
    return None


def compute_tobe(raw_pn: str, mfr: str):
    v = str(raw_pn).strip()
    removed_parts = []

    rule = LABEL_RULES.get(mfr)
    if rule:
        pattern, kind = rule
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()

    if mfr in HASH_STRIP_MFRS and v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if mfr in BRACKET_STRIP_MFRS and v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    if mfr == "Agilent":
        m = AGILENT_UI_RE.search(v)
        if m:
            removed_parts.append(v[m.start():])
            v = v[: m.start()]

    # 선두 '숫자+인치기호' 제거 (하이픈 접미사가 아닌 유일한 접두 규칙)
    m = _LEADING_INCH_RE.match(v)
    if m:
        removed_parts.append(m.group(0))
        v = v[m.end():].strip()

    suffix_split_done = False
    if mfr not in SUFFIX_ENGINE_EXCLUDE_MFRS:
        if not (mfr == "Thermo Fisher Scientific" and v.upper().startswith("KOLAS")):
            m = _SUFFIX_RE.match(v)
            if m:
                unit_up = m.group("unit").upper()
                matched = _unit_matches(unit_up)
                if matched:
                    suffix_text = v[len(m.group("core")):]
                    removed_parts.append(suffix_text)
                    v = m.group("core").strip()
                    suffix_split_done = True

    memo = ""
    if mfr in SIGMA_MFRS and not suffix_split_done:
        memo = "핵심코드 분리 안 됨(수량단위 패턴 불일치) - 원본 트림값 유지"

    separated_text = "".join(removed_parts)
    return v, memo, separated_text


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn) or pd.isna(mfr) or mfr not in TOP20_MFRS:
            continue
        tobe, memo, sep = compute_tobe(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시(Sigma 계열 미분리): {n_memo}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
