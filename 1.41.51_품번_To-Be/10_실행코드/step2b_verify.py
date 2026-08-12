# -*- coding: utf-8 -*-
"""[검증 전용] step2 산출물의 건수를 제조사별로 재계산해서 교차검증."""
import openpyxl
from collections import Counter

p = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260811_S-TEPS_품번_To-Be_v1.0.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
ws = wb["Steps_중복제거_32359"]

COL_MFR = 21   # 삽입 후: 원래 20 -> +1
COL_TOBE = 17
COL_MEMO = 18

mfr_total = Counter()
mfr_filled = Counter()
mfr_memo = Counter()

for r in range(5, ws.max_row + 1):
    mfr = ws.cell(row=r, column=COL_MFR).value
    tobe = ws.cell(row=r, column=COL_TOBE).value
    memo = ws.cell(row=r, column=COL_MEMO).value
    if tobe is not None:
        mfr_filled[mfr] += 1
        if memo:
            mfr_memo[mfr] += 1

print("제조사별 To-Be 채움 건수 / 비고 건수:")
for mfr, cnt in mfr_filled.most_common():
    print(f"  {mfr}: 채움 {cnt}건 / 비고 {mfr_memo.get(mfr,0)}건")

print()
print("합계 채움:", sum(mfr_filled.values()))
print("합계 비고:", sum(mfr_memo.values()))
print()
print("Sigma+Sigma-Aldrich 중 비고 없는(=핵심코드 분리 성공) 건수:",
      mfr_filled.get('Sigma',0) - mfr_memo.get('Sigma',0)
      + mfr_filled.get('Sigma-Aldrich',0) - mfr_memo.get('Sigma-Aldrich',0))
