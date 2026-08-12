# -*- coding: utf-8 -*-
"""[검증 전용] v2.0(원본유지) 파일의 열 밀림/데이터 무결성 검증."""
import openpyxl

SRC = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx"
DST = r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\20_결과\260811_S-TEPS_품번_To-Be_v2.0(원본유지).xlsx"

wb_src = openpyxl.load_workbook(SRC, data_only=False)
wb_dst = openpyxl.load_workbook(DST, data_only=False)

print("=== 시트 목록 비교 ===")
print("원본:", wb_src.sheetnames)
print("결과:", wb_dst.sheetnames)

ws_s = wb_src["Steps_중복제거_32359"]
ws_d = wb_dst["Steps_중복제거_32359"]

print()
print("=== 헤더행(4행) 비교: 원본 컬럼 -> 결과 컬럼(2칸 밀림 확인) ===")
for c in range(13, 24):
    print(f"  원본 col{c}({ws_s.cell(row=4,column=c).value!r})", end="  ")
    dst_c = c if c <= 16 else c + 2
    print(f"-> 결과 col{dst_c}({ws_d.cell(row=4,column=dst_c).value!r})")

print()
print("=== 새 컬럼(17,18) 헤더 ===")
print(" col17:", ws_d.cell(row=4, column=17).value)
print(" col18:", ws_d.cell(row=4, column=18).value)

print()
print("=== 데이터 무결성 샘플(5개 행, 원본 O~U열 vs 결과 O,P,새17,18,밀린19~23) ===")
import random
random.seed(1)
rows = random.sample(range(5, ws_s.max_row + 1), 5)
for r in rows:
    orig_vals = [ws_s.cell(row=r, column=c).value for c in range(13, 22)]
    dst_vals_before = [ws_d.cell(row=r, column=c).value for c in (13, 14, 15, 16)]
    dst_new = [ws_d.cell(row=r, column=c).value for c in (17, 18)]
    dst_vals_after = [ws_d.cell(row=r, column=c).value for c in range(19, 24)]
    print(f"row{r}")
    print("   원본(M~U):", orig_vals)
    print("   결과(M~P):", dst_vals_before, " | 신규(Q,R):", dst_new, " | 결과(S~W, 원래T~X):", dst_vals_after)

print()
print("=== M2 수식 확인 ===")
print(" 결과 M2:", ws_d["M2"].value)

print()
print("=== 다른 시트(예: erp_MM) 무손상 확인 ===")
ws_s2 = wb_src["erp_MM"]
ws_d2 = wb_dst["erp_MM"]
diffs = 0
for r in range(1, min(50, ws_s2.max_row + 1)):
    for c in range(1, ws_s2.max_column + 1):
        if ws_s2.cell(row=r, column=c).value != ws_d2.cell(row=r, column=c).value:
            diffs += 1
print(f" erp_MM 상위 50행 비교 중 차이 발견: {diffs}건")

print()
print("=== 제조사정리(신규 col22) 상위 20 재검증 ===")
from collections import Counter
c = Counter()
for r in range(5, ws_d.max_row + 1):
    tobe = ws_d.cell(row=r, column=17).value
    if tobe is not None:
        mfr = ws_d.cell(row=r, column=22).value
        c[mfr] += 1
for mfr, n in c.most_common(25):
    print(" ", mfr, n)
print(" 합계:", sum(c.values()))
