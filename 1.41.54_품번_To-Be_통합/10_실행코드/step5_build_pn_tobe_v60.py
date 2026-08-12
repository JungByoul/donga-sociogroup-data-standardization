# -*- coding: utf-8 -*-
r"""
v6.0: v5.1을 기반으로, 1.41.44(4차판정, 0810 스냅샷) 파일의 판정 컬럼 4개
(O열_1차판정/2차판정/3차판정/4차판정, '검토메모'는 제외)를 가져와 붙임.

두 파일은 행 순서가 서로 다르므로(정렬이 바뀜), A열 'key' 값으로 매칭함
(확인 결과 품번 있는 행 14,012/14,012 100% 매칭됨. key 중복 21건은 미미해 마지막 값 사용).

4차판정에서 이미 '진짜 품번 아님'으로 판정된 행(1차판정이 [품번]이 아니거나, 2~4차판정 중 하나라도
값이 있는 행 - 총 약 350건)은:
  - 품번_To-Be를 비움(사용자 확인 완료)
  - 품번_To-Be_분리텍스트에 원본 품번을 그대로 옮김
  - 품번_To-Be_의견은 v5.1 자체 로직 대신, 4차판정 사유를 x>형식으로 표시
    ([값없음]->x>값없음, [설명문키워드]->x>설명문키워드, [CAS의심]->x>CAS의심,
     [품명일치]->x>품목명동일, [숫자없음]->x>숫자없음, [단위값]->x>단위값, [구조식의심]->x>구조식의심)
그 외 나머지(약 13,663건)는 v5.1 로직 그대로.
"""
import re
import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from step3_build_pn_tobe_v51 import compute_tobe_routed  # noqa: E402

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
JUDGE_SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.44_품번_4차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v2.3.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.54_품번_To-Be_통합\20_결과\260812_S-TEPS_품번_To-Be_v6.0(4차판정컬럼추가).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_MFR_CLEAN_OLD = 20
COL_KEY = 1

INSERT_AT = 17
N_NEW_COLS = 3     # 품번_To-Be / 품번_To-Be_의견 / 품번_To-Be_분리텍스트 (v5.1과 동일)
JUDGE_INSERT_AT = 20  # v5.1 3개 컬럼(17,18,19) 바로 뒤
N_JUDGE_COLS = 4

# JUDGE_SRC 판정 컬럼 위치(1-based)
JCOL_1 = 16
JCOL_2 = 17
JCOL_3 = 18
JCOL_4 = 19

REASON_MAP = {
    "[값없음]": "값없음",
    "[설명문키워드]": "설명문키워드",
    "[CAS의심]": "CAS의심",
    "[품명일치]": "품목명동일",
    "[숫자없음]": "숫자없음",
    "[단위값]": "단위값",
    "[구조식의심]": "구조식의심",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def load_judge_map():
    """key -> (1차,2차,3차,4차) 판정값. 4개 컬럼 값을 그대로 가져옴."""
    wb = openpyxl.load_workbook(JUDGE_SRC, data_only=False)
    ws = wb[SHEET]
    m = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        key = row[COL_KEY - 1].value
        if key is None:
            continue
        j1 = row[JCOL_1 - 1].value
        j2 = row[JCOL_2 - 1].value
        j3 = row[JCOL_3 - 1].value
        j4 = row[JCOL_4 - 1].value
        m[key] = (j1, j2, j3, j4)
    return m


def reject_reason(j1, j2, j3, j4):
    """4차판정 결과에서 '이미 품번 아님' 사유를 뽑음. 없으면 None."""
    if j2 in REASON_MAP:
        return REASON_MAP[j2]
    if j3 in REASON_MAP:
        return REASON_MAP[j3]
    if j4 in REASON_MAP:
        return REASON_MAP[j4]
    if j1 is not None and j1 != "[품번]" and j1 in REASON_MAP:
        return REASON_MAP[j1]
    return None


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    print("[1/4] 4차판정 매핑 로드 중...")
    judge_map = load_judge_map()
    print(f"  key 매핑 {len(judge_map):,}건 로드")

    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[2/4] 복사: {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    # 7개 컬럼(v5.1 3개 + 판정 4개)을 한 번에 삽입 - 나눠서 삽입하면 위치가 밀려서 꼬임
    total_new = N_NEW_COLS + N_JUDGE_COLS
    ws.insert_cols(INSERT_AT, total_new)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, total_new)
        ws[coord] = fixed

    headers = [
        "품번_To-Be", "품번_To-Be_의견", "품번_To-Be_분리텍스트",
        "O열_1차판정", "O열_2차판정", "O열_3차판정", "O열_4차판정",
    ]
    widths = [22, 20, 20, 14, 14, 14, 14]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    print("[3/4] 값 채우는 중...")
    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]
    col_key_name = df.columns[COL_KEY - 1]

    n_filled = 0
    n_reject = 0
    n_keep = 0
    n_change = 0
    n_sep = 0
    n_key_missing = 0

    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        key = row[col_key_name]
        if pd.isna(pn):
            continue

        j1 = j2 = j3 = j4 = None
        if key in judge_map:
            j1, j2, j3, j4 = judge_map[key]
        else:
            n_key_missing += 1

        reason = reject_reason(j1, j2, j3, j4)

        if reason:
            tobe = ""
            sep = str(pn).strip()
            opinion = f"x>{reason}"
            n_reject += 1
            n_change += 1
        else:
            tobe, sep, opinion = compute_tobe_routed(pn, mfr)
            if opinion == "o>유지":
                n_keep += 1
            else:
                n_change += 1

        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        ws.cell(row=excel_row, column=INSERT_AT + 1, value=opinion)
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1
        ws.cell(row=excel_row, column=JUDGE_INSERT_AT, value=j1)
        ws.cell(row=excel_row, column=JUDGE_INSERT_AT + 1, value=j2)
        ws.cell(row=excel_row, column=JUDGE_INSERT_AT + 2, value=j3)
        ws.cell(row=excel_row, column=JUDGE_INSERT_AT + 3, value=j4)
        n_filled += 1

    print("[4/4] 저장 중...")
    wb.save(DST)
    print(f"[요약] 품번_To-Be 대상 행: {n_filled:,}건")
    print(f"[요약] 4차판정에서 이미 품번 아님으로 비운 행: {n_reject:,}건")
    print(f"[요약] o>유지: {n_keep:,}건 / x>변경(4차판정 포함): {n_change:,}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep:,}건")
    print(f"[요약] key 매핑 실패(4차판정 정보 없음): {n_key_missing:,}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
