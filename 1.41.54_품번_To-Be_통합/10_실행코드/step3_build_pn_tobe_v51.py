# -*- coding: utf-8 -*-
r"""
v5.1: v5.0(통합 결과)과 값 자체는 동일하되(버그 수정 1건 제외), '품번_To-Be_비고' 컬럼을
'품번_To-Be_의견'으로 이름을 바꾸고, 상사가 예전에 만든 '품번_의견' 컬럼과 같은 스타일
(o>유지 / x>사유)로 다시 씀.

의견 카테고리(6개, 사용자 확인 완료):
  o>유지            - 원본 그대로(공백 트림 외 아무것도 안 바뀜)
  x>수량단위 제거    - 수량+단위 접미사, 복합단위, NxM곱셈, 선두 인치기호, '-숫자중' 등
  x>안내라벨 제거    - 견적번호/부품번호/USP/시리얼번호/일련번호/Serial number/모델명/끝 외
  x>불용기호 제거    - 선두 '#', '[...]' 괄호, Agilent 'UI'
  x>부가정보 제거    - 끝 콤마+텍스트, '_Unstained', 개별 케이스 다수(부속설명/부가스펙 문구)
  x>상세규격 이동    - 밸브/글러브 규격문자열, 치수(0.1*420*300*500), 용기규격(LT/LC, 20L 용기 등)
  x>의미없음         - 엑셀론패취, 라인 레드 MXLR 등 코드로 볼 수 없는 서술형 값
한 행에서 규칙이 2개 이상 겹치면 'x>안내라벨+수량단위 제거'처럼 합쳐서 표기.

이번에 함께 고친 버그: 1.41.53의 v4.1 보고서는 '숫자+LT/LC 전체가 전체 제거된다'고 썼지만 실제
코드에는 그 처리가 빠져있었음(GASTEC '4LT'×3, '1LC'×2, '2LC'×2, 총 7건이 원본 그대로 남아있었음).
이번 버전에서 실제로 구현해서 반영함.

값(품번_To-Be, 분리텍스트) 자체는 이 LT/LC 수정 1건을 빼면 v5.0과 완전히 동일함 - 원본 3개 폴더의
compute_tobe 로직을 그대로 재사용하고, 여기에 '어떤 규칙이 발동했는지' 태그만 추가로 붙였음.
"""
import re
import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

sys.path.insert(0, r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\10_실행코드")
sys.path.insert(0, r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.52_품번_To-Be_차상위40\10_실행코드")
sys.path.insert(0, r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.53_품번_To-Be_나머지전체\10_실행코드")

import step8_build_pn_tobe_v23 as m20    # noqa: E402
import step5_build_pn_tobe_v32 as m40    # noqa: E402
import step3_build_pn_tobe_v41 as mrest  # noqa: E402

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.54_품번_To-Be_통합\20_결과\260812_S-TEPS_품번_To-Be_v5.1(의견컬럼_LTLC버그수정).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

TOP20_MFRS = m20.TOP20_MFRS
NEW40_MFRS = m40.NEW40_MFRS

_WHOLE_LT_LC_RE = re.compile(r"^\d+(\.\d+)?\s*(LT|LC)$", re.IGNORECASE)

# 개별 오버라이드 값 -> 의견 카테고리
OVERRIDE_CATEGORY = {
    "1203-0004-0048-03 for CV60-35-20 sensor": "부가정보",
    "NF-GM80-일반 길이조절 튜브": "부가정보",
    "NF-GM80-일반 노즐": "부가정보",
    "NF-GM80-일반 호스": "부가정보",
    "650 25D 88 40 5M 1 2T1 1508": "상세규격",
    "9550 15Z 5 1 1G1": "상세규격",
    "9554 15Z 5 1 1": "상세규격",
    "97800 Y 9 E4 5": "상세규격",
    "17877 UPN": "부가정보",
    "3S-Cath+ 24G": "수량단위",
    "HBP-8000CART-S": "부가정보",
    "KACON.K22-BZ 90~240V": "부가정보",
    "SW-Fitting-LB1/4in": "수량단위",
    "USP-76.55X4.40- 2107 (Custom)": "부가정보",
    "GMC-30P2(12300057)단상AC2": "부가정보",
    "EML 200 Premium": "부가정보",
}
BLANK_CATEGORY = {
    "20L 용기 깔대기옵션- 카본 필터": "상세규격",
    "달성-QC-이화학": "의미없음",
    "라인 레드 MXLR": "의미없음",
    "엑셀론패취10(리바스티그민)": "의미없음",
    "엑셀론패취15(리바스티그민)": "의미없음",
    "엑셀론패취5(리바스티그민)": "의미없음",
    "0.1*420*300*500": "상세규격",
    "0.1*430*300*500": "상세규격",
    "유포/무지/RP48": "의미없음",
    "지니언스3/8": "의미없음",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _cat_order(cats):
    order = ["안내라벨", "불용기호", "수량단위", "부가정보", "상세규격", "의미없음"]
    seen = [c for c in order if c in cats]
    return "+".join(seen)


def compute_top20_tagged(raw_pn, mfr):
    """1.41.51 step8(v2.3) 로직을 그대로 재현하면서 카테고리 태그를 붙임."""
    v = str(raw_pn).strip()
    cats = []

    rule = m20.LABEL_RULES.get(mfr)
    if rule:
        pattern, kind = rule
        mm = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if mm and mm.group(0):
            v = pattern.sub("", v).strip()
            cats.append("안내라벨")

    if mfr in m20.HASH_STRIP_MFRS and v.startswith("#"):
        v = v[1:].strip()
        cats.append("불용기호")

    if mfr in m20.BRACKET_STRIP_MFRS and v.startswith("[") and v.endswith("]") and len(v) >= 2:
        v = v[1:-1].strip()
        cats.append("불용기호")

    if mfr == "Agilent":
        mm = m20.AGILENT_UI_RE.search(v)
        if mm:
            v = v[: mm.start()]
            cats.append("불용기호")

    mm = m20._LEADING_INCH_RE.match(v)
    if mm:
        v = v[mm.end():].strip()
        cats.append("수량단위")

    if mfr not in m20.SUFFIX_ENGINE_EXCLUDE_MFRS:
        if not (mfr == "Thermo Fisher Scientific" and v.upper().startswith("KOLAS")):
            mm = m20._SUFFIX_RE.match(v)
            if mm:
                unit_up = mm.group("unit").upper()
                if m20._unit_matches(unit_up):
                    v = mm.group("core").strip()
                    cats.append("수량단위")

    tobe_real, memo_real, sep_real = m20.compute_tobe(raw_pn, mfr)
    assert v == tobe_real, f"태그 재현 불일치(top20): {raw_pn!r} tag={v!r} real={tobe_real!r}"

    opinion = "o>유지" if not cats else "x>" + _cat_order(cats) + " 제거"
    return tobe_real, sep_real, opinion


def compute_new40_tagged(raw_pn, mfr):
    """1.41.52 step5(v3.2) 로직을 그대로 재현하면서 카테고리 태그를 붙임."""
    v = str(raw_pn).strip()
    cats = []

    for pattern, kind in m40.GLOBAL_LABEL_RULES:
        mm = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if mm and mm.group(0):
            v = pattern.sub("", v).strip()
            cats.append("안내라벨")
            break

    if v.startswith("#"):
        v = v[1:].strip()
        cats.append("불용기호")

    if v.startswith("[") and v.endswith("]") and len(v) >= 2:
        v = v[1:-1].strip()
        cats.append("불용기호")

    mm = m40.GLOBAL_UI_RE.search(v)
    if mm:
        v = v[: mm.start()]
        cats.append("불용기호")

    mm = m40._LEADING_INCH_RE.match(v)
    if mm:
        v = v[mm.end():].strip()
        cats.append("수량단위")

    mm = m40._SUFFIX_RE.match(v)
    if mm:
        unit_up = mm.group("unit").upper()
        if m40._unit_matches(unit_up):
            v = mm.group("core").strip()
            cats.append("수량단위")

    tobe_real, memo_real, sep_real = m40.compute_tobe(raw_pn, mfr)
    assert v == tobe_real, f"태그 재현 불일치(new40): {raw_pn!r} tag={v!r} real={tobe_real!r}"

    opinion = "o>유지" if not cats else "x>" + _cat_order(cats) + " 제거"
    return tobe_real, sep_real, opinion


def compute_rest_tagged(raw_pn):
    """1.41.53 step3(v4.1) 로직을 재현 + LT/LC 전체제거 버그를 이번에 실제로 반영."""
    raw_stripped = str(raw_pn).strip()

    if raw_stripped in mrest.BLANK_ENTIRELY:
        cat = BLANK_CATEGORY[raw_stripped]
        return "", raw_stripped, f"x>{cat} 이동" if cat == "상세규격" else f"x>{cat}"

    if _WHOLE_LT_LC_RE.match(raw_stripped):
        return "", raw_stripped, "x>상세규격 이동"

    if raw_stripped in mrest.INDIVIDUAL_SUFFIX_OVERRIDES:
        tobe, sep = mrest.INDIVIDUAL_SUFFIX_OVERRIDES[raw_stripped]
        cat = OVERRIDE_CATEGORY[raw_stripped]
        verb = "이동" if cat == "상세규격" else "제거"
        return tobe, sep, f"x>{cat} {verb}"

    v = raw_stripped
    cats = []

    for pattern, kind in mrest.GLOBAL_LABEL_RULES:
        mm = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if mm and mm.group(0):
            v = pattern.sub("", v).strip()
            cats.append("안내라벨")
            break

    if v.startswith("#"):
        v = v[1:].strip()
        cats.append("불용기호")

    if v.startswith("[") and v.endswith("]") and len(v) >= 2:
        v = v[1:-1].strip()
        cats.append("불용기호")

    mm = mrest.GLOBAL_UI_RE.search(v)
    if mm:
        v = v[: mm.start()]
        cats.append("불용기호")

    mm = mrest._LEADING_INCH_RE.match(v)
    if mm:
        v = v[mm.end():].strip()
        cats.append("수량단위")

    mm = mrest._TRAILING_UNSTAINED_RE.search(v)
    if mm:
        v = v[: mm.start()]
        cats.append("부가정보")

    mm = mrest._TRAILING_JUNG_RE.match(v)
    if mm:
        v = mm.group("core").strip()
        cats.append("수량단위")

    mm = mrest._SUFFIX_RE.match(v)
    if mm:
        unit_up = mm.group("unit").upper()
        if mrest._unit_matches(unit_up):
            v = mm.group("core").strip()
            cats.append("수량단위")

    mm = mrest._TRAILING_COMMA_RE.search(v)
    if mm:
        v = v[: mm.start()].strip()
        cats.append("부가정보")

    tobe_real, sep_real, _memo_real = mrest.compute_tobe(raw_pn)
    assert v == tobe_real, f"태그 재현 불일치(rest): {raw_pn!r} tag={v!r} real={tobe_real!r}"

    opinion = "o>유지" if not cats else "x>" + _cat_order(cats) + " 제거"
    return tobe_real, sep_real, opinion


def compute_tobe_routed(raw_pn, mfr):
    if mfr in TOP20_MFRS:
        return compute_top20_tagged(raw_pn, mfr)
    elif mfr in NEW40_MFRS:
        return compute_new40_tagged(raw_pn, mfr)
    else:
        return compute_rest_tagged(raw_pn)


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed

    headers = ["품번_To-Be", "품번_To-Be_의견", "품번_To-Be_분리텍스트"]
    widths = [22, 20, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_keep = 0
    n_change = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn):
            continue
        tobe, sep, opinion = compute_tobe_routed(pn, mfr)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        ws.cell(row=excel_row, column=INSERT_AT + 1, value=opinion)
        n_filled += 1
        if opinion == "o>유지":
            n_keep += 1
        else:
            n_change += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] o>유지: {n_keep}건 / x>변경: {n_change}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
