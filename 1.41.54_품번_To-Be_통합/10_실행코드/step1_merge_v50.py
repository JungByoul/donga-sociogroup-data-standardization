# -*- coding: utf-8 -*-
r"""
v5.0: 상위20(1.41.51 v2.3) + 차상위40(1.41.52 v3.2) + 나머지전체(1.41.53 v4.1) 3개 결과를
1개 파일로 통합.

각 파일이 서로 다른(안 겹치는) 회사 범위를 담당했으므로, 행마다 소속 범위를 판별해서 해당 범위의
compute_tobe 로직을 그대로 재실행하는 방식으로 통합함(3개 파일을 직접 읽어붙이는 대신, 이미 검증된
로직을 원본에서 재실행 - 지금까지의 작업 방식과 동일).

제조사(정리) 기준 판별:
- 상위20(TOP20_MFRS) 이면 -> 1.41.51 step8(v2.3) 로직
- 차상위40(NEW40_MFRS) 이면 -> 1.41.52 step5(v3.2) 로직
- 그 외(품번 있고 위 60개가 아닌 모든 행, 제조사 미기재 포함) -> 1.41.53 step3(v4.1) 로직

주의: '제조사_정리' 값은 1글자라도 다르면 다른 회사로 취급(예: 'Agilent'와 'Agilent Technologies'는
별개 회사) - 정성적 병합 없음.
"""
import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter
import re

sys.path.insert(0, r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.51_품번_To-Be\10_실행코드")
sys.path.insert(0, r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.52_품번_To-Be_차상위40\10_실행코드")
sys.path.insert(0, r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.53_품번_To-Be_나머지전체\10_실행코드")

import step8_build_pn_tobe_v23 as m20   # noqa: E402  (compute_tobe(raw, mfr) -> tobe, memo, sep)
import step5_build_pn_tobe_v32 as m40   # noqa: E402  (compute_tobe(raw, mfr) -> tobe, memo, sep)
import step3_build_pn_tobe_v41 as mrest  # noqa: E402  (compute_tobe(raw) -> tobe, sep, memo)

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.54_품번_To-Be_통합\20_결과\260812_S-TEPS_품번_To-Be_v5.0(통합_상위20_차상위40_나머지).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

TOP20_MFRS = m20.TOP20_MFRS
NEW40_MFRS = m40.NEW40_MFRS

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def compute_tobe_routed(raw_pn, mfr):
    if mfr in TOP20_MFRS:
        tobe, memo, sep = m20.compute_tobe(raw_pn, mfr)
        return tobe, memo, sep, "상위20(v2.3)"
    elif mfr in NEW40_MFRS:
        tobe, memo, sep = m40.compute_tobe(raw_pn, mfr)
        return tobe, memo, sep, "차상위40(v3.2)"
    else:
        tobe, sep, memo = mrest.compute_tobe(raw_pn)
        return tobe, memo, sep, "나머지전체(v4.1)"


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_memo = 0
    n_sep = 0
    src_counter = {"상위20(v2.3)": 0, "차상위40(v3.2)": 0, "나머지전체(v4.1)": 0}
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn):
            continue
        tobe, memo, sep, source = compute_tobe_routed(pn, mfr)
        src_counter[source] += 1
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_memo += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 비고 표시: {n_memo}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[요약] 출처별 건수: {src_counter}")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
