# -*- coding: utf-8 -*-
r"""
중간수정: 원본 손상 복구 + 단계별 판정 열 분리

문제:
1) v0.4에서 사용자가 Q열(CAS No., 원본 그대로여야 함)에 "원본 -> 보정값"을 직접 입력해
   원본 데이터가 훼손됨.
2) CAS_1차판정 한 열에 1차/2차/3차 결과가 계속 덮어써져서 단계별 이력이 안 보임.

해결:
- Q열은 손대지 않은 진짜 원본 파일((input_원본)...v.0.5.xlsx)에서 그대로 가져옴.
- 1차 결과는 1단계 산출물(v.0.2, [CAS아님] 라벨까지만 반영된 상태)에서 그대로 가져옴.
- 2차 결과(수기 보정 9건)는 3단계 산출물(v.0.5)의 "... [오타수정함]" 값에서 추출.
- 3차 결과([미상]/[화학식]/[품번/문서번호])도 v.0.5에서 그대로 가져옴.
- Q~R 사이에 CAS_1차판정 / 품번대조경고 / CAS_2차판정 / CAS_3차판정 4개 열을 삽입.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

BASE = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리")

ORIGINAL = BASE / "(input_원본)260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5.xlsx"
STAGE1_V02 = BASE / "10_CAS_1단계" / "20_결과" / "260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.2.xlsx"
STAGE3_V05 = BASE / "30_CAS_3단계" / "20_결과" / "260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5.xlsx"
DST = BASE / "40_중간수정" / "20_결과" / "260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.6.xlsx"

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5

COL_Q_ORIG = 17          # 원본 파일 Q열 (CAS No.)
COL_CAS1_V02 = 18        # v0.2의 R열 (CAS_1차판정: 값 또는 [CAS아님])
COL_WARN_V02 = 19        # v0.2의 S열 (품번대조경고)
COL_RESULT_V05 = 18      # v0.5의 R열 (1/2/3차 결과가 섞여 있는 열)

INSERT_AT = 18           # Q(17) 다음, 원본 R열 앞
N_NEW_COLS = 4           # CAS_1차판정 / 품번대조경고 / CAS_2차판정 / CAS_3차판정

_TYPO_FIX_RE = re.compile(r"^(.*) \[오타수정함\]$")
STAGE3_LABELS = {"[미상]", "[화학식]", "[품번/문서번호]"}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb_orig = openpyxl.load_workbook(ORIGINAL, data_only=False)
    ws_orig = wb_orig[SHEET]

    wb_v02 = openpyxl.load_workbook(STAGE1_V02, data_only=False)
    ws_v02 = wb_v02[SHEET]

    wb_v05 = openpyxl.load_workbook(STAGE3_V05, data_only=False)
    ws_v05 = wb_v05[SHEET]

    max_row = ws_orig.max_row
    assert max_row == ws_v02.max_row == ws_v05.max_row, "행 수 불일치 - 정렬/필터 여부 확인 필요"

    # 원본 워크북을 베이스로 사용해서 Q열 및 나머지 열은 자동으로 원본 그대로 유지됨
    ws = ws_orig

    computed = {}  # row -> (cas1, warn, cas2, cas3)
    n_valid1 = n_none1 = n_stage2 = n_stage3 = 0
    for r in range(DATA_START_ROW, max_row + 1):
        cas1 = ws_v02.cell(row=r, column=COL_CAS1_V02).value
        warn = ws_v02.cell(row=r, column=COL_WARN_V02).value
        result5 = ws_v05.cell(row=r, column=COL_RESULT_V05).value

        cas2 = ""
        cas3 = ""
        if isinstance(result5, str):
            m = _TYPO_FIX_RE.match(result5)
            if m:
                cas2 = result5  # "보정값 [오타수정함]" 그대로
                n_stage2 += 1
            elif result5 in STAGE3_LABELS:
                cas3 = result5
                n_stage3 += 1

        if cas1 not in (None, ""):
            if cas1 == "[CAS아님]":
                n_none1 += 1
            else:
                n_valid1 += 1

        if any([cas1, warn, cas2, cas3]):
            computed[r] = (cas1 or "", warn or "", cas2, cas3)

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    ws.cell(row=4, column=INSERT_AT, value="CAS_1차판정")
    ws.cell(row=4, column=INSERT_AT + 1, value="품번대조경고")
    ws.cell(row=4, column=INSERT_AT + 2, value="CAS_2차판정")
    ws.cell(row=4, column=INSERT_AT + 3, value="CAS_3차판정")

    for r, (cas1, warn, cas2, cas3) in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=cas1)
        ws.cell(row=r, column=INSERT_AT + 1, value=warn)
        ws.cell(row=r, column=INSERT_AT + 2, value=cas2)
        ws.cell(row=r, column=INSERT_AT + 3, value=cas3)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb_orig.save(DST)

    print(f"[검증] Q열은 원본 파일에서 그대로 가져옴 (수기 주석 없음)")
    print(f"[요약] CAS_1차판정 유효: {n_valid1} / [CAS아님]: {n_none1}")
    print(f"[요약] CAS_2차판정(수기보정) 채워진 행: {n_stage2}")
    print(f"[요약] CAS_3차판정 채워진 행: {n_stage3}")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
