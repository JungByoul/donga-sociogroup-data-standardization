# -*- coding: utf-8 -*-
r"""
v0.4(사용자 수기 9건 입력) -> v0.5 (3단계)

1) 2단계 마무리: [체크섬오류] 9건은 사용자가 Q열에 "원본 -> 보정값" 형태로 수기 입력해둠.
   -> 보정값을 CAS_1차판정에 반영해 [체크섬오류] 라벨을 해소.
2) 3단계: 남은 [CAS아님] 317건을 전수 검토(수작업 분류, 품번 정형 패턴이 없어 규칙화 불가)해서
   - 진짜 값없음/설명텍스트(플레이스홀더): [미상]
   - 분자식(화학식이 잘못 들어간 경우): [화학식]
   - 그 외 카탈로그/로트/문서번호 형태로 판단되는 나머지: [품번/문서번호]
   로 최종 라벨링.
"""
import re
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\20_CAS_2단계\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.4_수기_9건입력.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\30_CAS_3단계\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_Q_CAS원본 = 17  # Q
COL_CAS_1차판정 = 18  # R

LABEL_NONE = "[CAS아님]"
LABEL_CHECKSUM_ERR = "[체크섬오류]"
LABEL_UNKNOWN = "[미상]"
LABEL_FORMULA = "[화학식]"
LABEL_ITEM_NO = "[품번/문서번호]"

# 전수 검토(317건) 결과 - 값없음/설명텍스트 플레이스홀더
UNKNOWN_VALUES = {"1", "-", "없음", "N//A", "화이트", "Male, 7주령"}
# 전수 검토 결과 - 분자식(화학식)이 CAS No. 칸에 잘못 들어간 값
FORMULA_VALUES = {"C17H19D3N2O2", "C17H22N2O2"}

_ARROW_RE = re.compile(r"->\s*(\S+)\s*$")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    fixed_checksum_err = 0
    unknown_cnt = 0
    formula_cnt = 0
    item_no_cnt = 0

    for r in range(DATA_START_ROW, ws.max_row + 1):
        cas1_cell = ws.cell(row=r, column=COL_CAS_1차판정)
        q_val = ws.cell(row=r, column=COL_Q_CAS원본).value
        q_str = "" if q_val is None else str(q_val).strip()

        if cas1_cell.value == LABEL_CHECKSUM_ERR:
            m = _ARROW_RE.search(q_str)
            if m:
                cas1_cell.value = f"{m.group(1)} [오타수정함]"
                fixed_checksum_err += 1
            continue

        if cas1_cell.value != LABEL_NONE:
            continue

        if q_str in FORMULA_VALUES:
            cas1_cell.value = LABEL_FORMULA
            formula_cnt += 1
        elif q_str in UNKNOWN_VALUES:
            cas1_cell.value = LABEL_UNKNOWN
            unknown_cnt += 1
        else:
            cas1_cell.value = LABEL_ITEM_NO
            item_no_cnt += 1

    wb.save(DST)
    print(f"[요약] 체크섬오류 -> 수기보정값 반영: {fixed_checksum_err}건")
    print(f"[요약] {LABEL_UNKNOWN}: {unknown_cnt}건")
    print(f"[요약] {LABEL_FORMULA}: {formula_cnt}건")
    print(f"[요약] {LABEL_ITEM_NO}: {item_no_cnt}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
