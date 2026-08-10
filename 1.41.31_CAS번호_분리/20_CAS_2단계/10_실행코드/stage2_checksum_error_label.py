# -*- coding: utf-8 -*-
r"""
v0.2 -> v0.3 (2단계)

2단계 결론(데이터 실측 검증 결과):
- Q열이 대시 없는 순수 숫자(5~10자리)인 39건에 대해 CAS 표준 자리수 규칙
  (끝 1자리=체크디지트, 그 앞 2자리=중간그룹, 나머지=앞그룹)으로 대시를 복원해 체크디지트
  검증을 시도했으나, 39건 중 2건만 통과했고 그 2건도 품목명이 항체/바이얼(화학물질 아님)이며
  품번과 CAS No. 값이 완전히 동일 -> 품번이 잘못 복붙된 것으로 판단, 우연한 체크섬 일치.
  => 39건 전부 [CAS아님] 그대로 유지 (자동 복원 없음)
- Q열이 이미 CAS 형태(\d{2,7}-\d{2}-\d)로 대시가 정상 위치인데 체크디지트만 실패한 9건은
  "대시 누락"이 아니라 숫자 자체의 오타 가능성 -> 자동 복구 대상에서 제외하고
  [체크섬오류] 라벨로 구분 표시만 함 (사람이 우선순위 높게 검토하도록)
"""
import re
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\10_CAS_1단계\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.2.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\20_CAS_2단계\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.3.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_Q_CAS원본 = 17  # Q
COL_CAS_1차판정 = 18  # R

_CAS_SHAPE_RE = re.compile(r"\b\d{2,7}-\d{2}-\d\b")
LABEL_NONE = "[CAS아님]"
LABEL_CHECKSUM_ERR = "[체크섬오류]"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    relabeled = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        cas1_cell = ws.cell(row=r, column=COL_CAS_1차판정)
        if cas1_cell.value != LABEL_NONE:
            continue
        q_val = ws.cell(row=r, column=COL_Q_CAS원본).value
        if q_val is None:
            continue
        if _CAS_SHAPE_RE.search(str(q_val)):
            cas1_cell.value = LABEL_CHECKSUM_ERR
            relabeled += 1

    wb.save(DST)
    print(f"[요약] '{LABEL_CHECKSUM_ERR}'로 재라벨링: {relabeled}행")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
