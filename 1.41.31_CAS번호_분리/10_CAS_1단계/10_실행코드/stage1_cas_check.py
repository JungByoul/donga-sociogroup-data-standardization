# -*- coding: utf-8 -*-
r"""
CAS No. 1차 판정
- Steps_중복제거_32359 시트, Q열(CAS No.) 값을 검사해서
  "실제 CAS번호 형태(\d{2,7}-\d{2}-\d) + 공식 체크디지트(mod-10)"를 통과하는 값을 추출.
- Q열과 R열 사이에 CAS_1차판정 / 품번대조경고 2개 컬럼을 삽입.
- 원본은 건드리지 않고 20_결과 폴더에 새 버전으로 저장.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\10_CAS_1단계\20_결과\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v0.6.xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_O_품번 = 15  # O
COL_Q_CAS = 17  # Q
INSERT_AT = 18  # R 앞에 삽입 (Q와 R 사이)
N_NEW_COLS = 2

# parse_sds_v6.3_실제값추출.py 의 _CAS_RE 재사용 (715행): r'\b(\d{2,7}-\d{2}-\d)\b'
_CAS_RE = re.compile(r"\b(\d{2,7}-\d{2}-\d)\b")

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def cas_checksum_valid(cas: str) -> bool:
    """CAS 공식 체크디지트(mod-10) 검증. [외부출처] 정확한 원문서는 특정 불가 —
    데이터 내 실존 CAS(58-08-2, 7732-18-5 등)로 CMD 검산해 확인함."""
    digits = cas.replace("-", "")
    body, check = digits[:-1], int(digits[-1])
    total = sum(int(d) * (i + 1) for i, d in enumerate(reversed(body)))
    return total % 10 == check


def extract_valid_cas(cell_value) -> str:
    """셀 안에서 CAS 형태 후보를 모두 찾아 체크디지트까지 통과한 값만 반환 ('; ' join)."""
    if cell_value is None:
        return ""
    text = str(cell_value)
    candidates = _CAS_RE.findall(text)
    valid = []
    for c in candidates:
        if cas_checksum_valid(c) and c not in valid:
            valid.append(c)
    return "; ".join(valid)


def item_no_warning(cas_hit: str, item_no_value) -> str:
    """1차판정 추출값과 O열(품번)이 숫자만 놓고 봤을 때 겹치면 경고."""
    if not cas_hit or item_no_value is None:
        return ""
    cas_digits = re.sub(r"\D", "", cas_hit)
    item_digits = re.sub(r"\D", "", str(item_no_value))
    if not cas_digits or not item_digits:
        return ""
    if cas_digits == item_digits or cas_digits in item_digits or item_digits in cas_digits:
        return "⚠ 품번(O열)과 유사"
    return ""


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    """수식 안 셀참조 중 insert_at 이상 컬럼을 n칸 오른쪽으로 보정."""

    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    # 사전 점검에서 확인된 유일한 수식(M2, AA열 참조)을 컬럼 삽입 전에 보정 대상으로 수집
    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))
    print(f"[사전점검] {SHEET} 시트 내 수식 {len(formula_cells)}개: {formula_cells}")

    total_rows = ws.max_row
    q_nonempty = 0
    cas_valid = 0
    warn_count = 0

    # 삽입 전에 원본 Q/O 값을 먼저 다 읽어둔다 (삽입 후에는 열 위치가 밀리므로)
    computed = {}  # row -> (cas_hit, warn)
    for r in range(DATA_START_ROW, total_rows + 1):
        q_val = ws.cell(row=r, column=COL_Q_CAS).value
        o_val = ws.cell(row=r, column=COL_O_품번).value
        if q_val not in (None, ""):
            q_nonempty += 1
        cas_hit = extract_valid_cas(q_val)
        warn = item_no_warning(cas_hit, o_val)
        if cas_hit:
            cas_valid += 1
        if warn:
            warn_count += 1
        if cas_hit or warn:
            computed[r] = (cas_hit, warn)

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    # 수식 보정 (삽입으로 인해 컬럼이 밀린 참조 갱신)
    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    ws.cell(row=HEADER_ROW, column=INSERT_AT, value="CAS_1차판정")
    ws.cell(row=HEADER_ROW, column=INSERT_AT + 1, value="품번대조경고")

    for r, (cas_hit, warn) in computed.items():
        ws.cell(row=r, column=INSERT_AT, value=cas_hit)
        ws.cell(row=r, column=INSERT_AT + 1, value=warn)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print(f"[요약] 전체 데이터행: {total_rows - DATA_START_ROW + 1}")
    print(f"[요약] Q열(CAS No.) 값 있는 행: {q_nonempty}")
    print(f"[요약] CAS_1차판정 유효(실제 CAS로 인정): {cas_valid}")
    print(f"[요약] Q열 값은 있으나 1차판정 공란(비CAS, 2차/3차로 이월): {q_nonempty - cas_valid}")
    print(f"[요약] 품번대조경고 발생: {warn_count}")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
