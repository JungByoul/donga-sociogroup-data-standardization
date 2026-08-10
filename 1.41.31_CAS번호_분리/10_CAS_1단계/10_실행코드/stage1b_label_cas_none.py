# -*- coding: utf-8 -*-
"""
v0.1 -> v0.2
모수는 Q열(CAS No.)에 값이 입력된 행(2,304건)만 해당.
그중 CAS_1차판정(R열)이 공란인 행에만 '[CAS아님]' 표지를 단다.
Q열 자체가 비어있던 행은 건드리지 않고 공란 유지.
"""
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\10_CAS_1단계\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.1.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.31_CAS번호_분리\10_CAS_1단계\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.2.xlsx")

SHEET = "Steps_중복제거_32359"
DATA_START_ROW = 5
COL_Q_CAS원본 = 17  # Q
COL_CAS_1차판정 = 18  # R
LABEL = "[CAS아님]"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    q_nonempty = 0
    labeled = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        q_val = ws.cell(row=r, column=COL_Q_CAS원본).value
        if q_val in (None, ""):
            continue  # Q열 자체가 공란인 행은 모수에서 제외 - 건드리지 않음
        q_nonempty += 1
        cell = ws.cell(row=r, column=COL_CAS_1차판정)
        if cell.value in (None, ""):
            cell.value = LABEL
            labeled += 1

    wb.save(DST)
    print(f"[요약] Q열 값 있는 행(모수): {q_nonempty}")
    print(f"[요약] '{LABEL}' 라벨 부여: {labeled}행")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
