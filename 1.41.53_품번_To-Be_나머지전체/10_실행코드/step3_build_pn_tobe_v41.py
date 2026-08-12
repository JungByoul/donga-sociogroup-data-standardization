# -*- coding: utf-8 -*-
r"""
v4.1: 6,722건 전수 수기검토(사용자) 결과 반영.

일반 규칙(전역, 재현 가능한 패턴 - 6개):
1) 끝에 콤마+텍스트가 붙은 경우, 마지막 콤마 이후 부분만 1회 제거.
   예: '#3662-4000, Nalgene' -> '3662-4000'(#는 기존 규칙으로 별도 제거),
       'DH.BOG011,대한과학' -> 'DH.BOG011'
2) 끝 '_Unstained' 제거 (TissueArray 14건).
3) 복합단위 'ML-R' 추가 (Fluka '34828-40ML-R' -> '34828').
4) '-숫자중'(한글 단위 '중') 제거 (Komed '500551-1중' -> '500551').
5) 수량단위 접미사 뒤에 '(브랜드명)'이 괄호로 더 붙은 경우, 그 괄호까지 함께 분리.
   예: 'Z373427-50EA(Merck)' -> 'Z373427'.
6) 전역 라벨 규칙에 'Serial number:' 추가(영문판, 기존 '시리얼 번호:'/'일련번호'와 별개 표현).
7) 값 전체가 '숫자+LT' 또는 '숫자+LC'인 경우(용기 규격 라벨로 판단, 코드 자체가 없음) 품번_To-Be를
   통째로 비움. 예: GASTEC '4LT', '1LC', '2LC'.

개별 처리(패턴 일반화 위험이 있어 실제 발견된 값에 대해서만 적용 - 사용자 확인 완료):
- 접미 분리 12건: Optek(for ... sensor), nilfisk 3건(-한글 설명), 밸브/글러브 규격문자열 4건,
  Sartorius Stedim(UPN), 덕우메디칼(+24G), nan(CART), nan(90~240V), 석림세이프티(in),
  Monucla(Custom), nan(단상AC2), NEXOPART(Premium)
- 전체 제거(품번_To-Be를 통째로 비움) 9건: 세이프랩, Mettler Toledo Korea(달성-QC-이화학),
  탑세이프티, 노바티스 3건(엑셀론패취), nan 2건(치수 규격), 버텍스아이디, 미래종합상사

주의: 개별 처리 항목은 '이 정확한 원본값'에 대해서만 적용되는 화이트리스트 방식이며, 비슷하게 생긴
다른 값에는 일반화되지 않음(예: -한글 접미사는 nilfisk 3건 외에는 그대로 유지 — '경인에스브이씨'의
색상/모델명으로 보이는 'UN-나인나인' 등은 건드리지 않음).
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.53_품번_To-Be_나머지전체\20_결과\260812_S-TEPS_품번_To-Be_v4.1(수기검토반영_나머지전체).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_PN_CLEAN = 16
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "RXN", "TAB", "CAP", "EA", "KT", "VL", "MM", "ML", "UL", "G", "L", "M", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)

COMPOUND_UNITS = {"AMP-EA", "KG-K", "G-F", "SET-F", "G-K", "ML-R"}

_NUM = r"(?:\d+(?:\.\d+)?|\.\d+)"
_QTY = rf"(?:{_NUM}[xX]{_NUM}|{_NUM})"
_UNIT = r"[A-Za-z%]+(?:-[A-Za-z%]+)?"
_SUFFIX_RE = re.compile(rf"^(?P<core>.+?)-\s*(?P<qty>{_QTY})\s*(?P<unit>{_UNIT})(?P<paren>\([^)]*\))?\Z")

_LEADING_INCH_RE = re.compile(rf"^(?P<qty>{_NUM})\s*\\?\"\s*")
_TRAILING_COMMA_RE = re.compile(r",\s*[^,]+\Z")
_TRAILING_UNSTAINED_RE = re.compile(r"_Unstained\Z")
_TRAILING_JUNG_RE = re.compile(r"^(?P<core>.+?)-\d+중\Z")
_WHOLE_LT_LC_RE = re.compile(r"^\d+(\.\d+)?\s*(LT|LC)$", re.IGNORECASE)

GLOBAL_LABEL_RULES = [
    (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^USP\s+"), "prefix"),
    (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^일련\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^Serial\s*number\s*[:：]?\s*", re.IGNORECASE), "prefix"),
    (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
    (re.compile(r"\s*외\s*\Z"), "suffix"),
]
GLOBAL_UI_RE = re.compile(r"UI\Z", re.IGNORECASE)

# 원본 그대로(str(raw).strip() 기준) -> (tobe, sep) 개별 지정
INDIVIDUAL_SUFFIX_OVERRIDES = {
    "1203-0004-0048-03 for CV60-35-20 sensor": ("1203-0004-0048-03", " for CV60-35-20 sensor"),
    "NF-GM80-일반 길이조절 튜브": ("NF-GM80", "-일반 길이조절 튜브"),
    "NF-GM80-일반 노즐": ("NF-GM80", "-일반 노즐"),
    "NF-GM80-일반 호스": ("NF-GM80", "-일반 호스"),
    "650 25D 88 40 5M 1 2T1 1508": ("650", " 25D 88 40 5M 1 2T1 1508"),
    "9550 15Z 5 1 1G1": ("9550", " 15Z 5 1 1G1"),
    "9554 15Z 5 1 1": ("9554", " 15Z 5 1 1"),
    "97800 Y 9 E4 5": ("97800", " Y 9 E4 5"),
    "17877 UPN": ("17877", " UPN"),
    "3S-Cath+ 24G": ("3S-Cath", "+ 24G"),
    "HBP-8000CART-S": ("HBP", "-8000CART-S"),
    "KACON.K22-BZ 90~240V": ("KACON.K22-BZ", " 90~240V"),
    "SW-Fitting-LB1/4in": ("SW-Fitting-LB1", "/4in"),
    "USP-76.55X4.40- 2107 (Custom)": ("USP-76.55X4.40- 2107", " (Custom)"),
    "GMC-30P2(12300057)단상AC2": ("GMC-30P2(12300057)", "단상AC2"),
    "EML 200 Premium": ("EML 200", " Premium"),
}

# 원본 그대로(str(raw).strip() 기준) -> 품번_To-Be 전체를 비움(서술형 텍스트/규격, 코드 아님)
BLANK_ENTIRELY = {
    "20L 용기 깔대기옵션- 카본 필터",
    "달성-QC-이화학",
    "라인 레드 MXLR",
    "엑셀론패취10(리바스티그민)",
    "엑셀론패취15(리바스티그민)",
    "엑셀론패취5(리바스티그민)",
    "0.1*420*300*500",
    "0.1*430*300*500",
    "유포/무지/RP48",
    "지니언스3/8",
}
BLANK_MEMO = "서술형 값/규격으로 판단되어 품번_To-Be에서 제외(수기검토)"

TOP60_MFRS = {
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
    "Charles River", "Km", "R&D Systems", "Combi-Blocks", "Abcam", "PerkinElmer", "대정화금",
    "BioLegend", "SPL", "Saint-Gobain", "Supelco", "싸토리우스코리아", "영인에스티", "TCI", "비티알",
    "DURAN", "Promega", "DAE", "EP", "Cobetter", "MedChemExpress", "ProteinSimple", "GEMÜ",
    "Lonza", "Metrohm", "Avantor", "Beckman Coulter", "Agilent Technologies", "(주)BB", "TRC",
    "Repligen", "Millipore", "Isolab", "Siemens", "Daejung", "Gilson", "DUKSAN", "Rhawn",
    "시너지이노베이션", "Advanced Instruments",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _unit_matches(unit_up: str):
    if unit_up in COMPOUND_UNITS:
        return unit_up
    if "-" in unit_up:
        return None
    if unit_up in UNIT_WHITELIST:
        return unit_up
    for wl in UNIT_WHITELIST:
        if unit_up.startswith(wl):
            return wl
    return None


def compute_tobe(raw_pn: str):
    raw_stripped = str(raw_pn).strip()

    if raw_stripped in BLANK_ENTIRELY:
        return "", raw_stripped, BLANK_MEMO

    if raw_stripped in INDIVIDUAL_SUFFIX_OVERRIDES:
        tobe, sep = INDIVIDUAL_SUFFIX_OVERRIDES[raw_stripped]
        return tobe, sep, ""

    v = raw_stripped
    removed_parts = []

    for pattern, kind in GLOBAL_LABEL_RULES:
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()
            break

    if v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    m = GLOBAL_UI_RE.search(v)
    if m:
        removed_parts.append(v[m.start():])
        v = v[: m.start()]

    m = _LEADING_INCH_RE.match(v)
    if m:
        removed_parts.append(m.group(0))
        v = v[m.end():].strip()

    m = _TRAILING_UNSTAINED_RE.search(v)
    if m:
        removed_parts.append(v[m.start():])
        v = v[: m.start()]

    m = _TRAILING_JUNG_RE.match(v)
    if m:
        removed_parts.append(v[len(m.group("core")):])
        v = m.group("core").strip()

    m = _SUFFIX_RE.match(v)
    if m:
        unit_up = m.group("unit").upper()
        matched = _unit_matches(unit_up)
        if matched:
            suffix_text = v[len(m.group("core")):]
            removed_parts.append(suffix_text)
            v = m.group("core").strip()

    # 콤마 트레일링(마지막 콤마 이후 1개 세그먼트)은 다른 규칙들 다음, 맨 마지막에 1회만 적용
    m = _TRAILING_COMMA_RE.search(v)
    if m:
        removed_parts.append(v[m.start():])
        v = v[: m.start()].strip()

    separated_text = "".join(removed_parts)
    return v, separated_text, ""


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_sep = 0
    n_blank = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn):
            continue
        if mfr in TOP60_MFRS:
            continue
        tobe, sep, memo = compute_tobe(pn)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if memo:
            ws.cell(row=excel_row, column=INSERT_AT + 1, value=memo)
            n_blank += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 전체 제거(비움) 행: {n_blank}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
