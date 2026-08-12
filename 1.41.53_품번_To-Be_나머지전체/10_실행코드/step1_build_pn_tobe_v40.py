# -*- coding: utf-8 -*-
r"""
v4.0: 상위20(1.41.51) + 차상위40(1.41.52)에서 완성된 '전역 규칙'(v3.2 기준)만 그대로 재사용해서,
나머지 전체 회사(제조사(정리) 61위 이하 + 제조사 미기재 포함)에 적용.

포함되는 전역 규칙(회사명 무관, 내용만 맞으면 적용):
- 라벨 프리픽스/서픽스 제거: 견적번호/부품 번호/USP/시리얼 번호/일련번호/모델명/끝 외
- 선두 '#' 제거
- 전체가 '[...]'로 감싸진 경우 제거
- 끝 'UI' 제거
- 선두 '숫자+인치기호' 제거
- 공통 수량단위 분리(화이트리스트: KG/MG/UG/NG/AMP/PAK/RXN/TAB/CAP/EA/KT/VL/MM/ML/UL/G/L/M/%,
  복합단위 AMP-EA/KG-K/G-F/SET-F/G-K, NxM 곱셈표기, 공백하이픈, 앞자리 없는 소수)

제외한 것(사용자 확인 완료 — 특정 3개 회사에서만 발견된 실제 위험 사례라 이번 범위에는 일반화하지 않음):
- 대한과학/Agilent 공통 수량단위 분리 스킵 예외 (해당 회사는 이미 상위20에서 처리 완료, 이번 범위에
  없음)
- Thermo Fisher Scientific 'KOLAS' 예외 (동일 이유로 이번 범위에 없음)

처리 범위: 품번이 있고, 제조사(정리)가 상위60(TOP20+NEW40)에 속하지 않는 모든 행
(제조사 미기재 행 포함). 상위60(7,291건, 이미 완료된 v2.3/v3.2)은 이번 범위에서 완전히 제외하고
건드리지 않음.
"""
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.50_작업중_파일_공유받음(0811)\260428_S-TEPS_입고실적만 ◆_최근3개년_uniq_v.0.5(0811_16시).xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.53_품번_To-Be_나머지전체\20_결과\260812_S-TEPS_품번_To-Be_v4.0(전역규칙_나머지전체).xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_PN = 15
COL_PN_CLEAN = 16
INSERT_AT = 17
N_NEW_COLS = 3
COL_MFR_CLEAN_OLD = 20

UNIT_WHITELIST = ["KG", "MG", "UG", "NG", "AMP", "PAK", "RXN", "TAB", "CAP", "EA", "KT", "VL", "MM", "ML", "UL", "G", "L", "M", "%"]
UNIT_WHITELIST.sort(key=len, reverse=True)

COMPOUND_UNITS = {"AMP-EA", "KG-K", "G-F", "SET-F", "G-K"}

_NUM = r"(?:\d+(?:\.\d+)?|\.\d+)"
_QTY = rf"(?:{_NUM}[xX]{_NUM}|{_NUM})"
_UNIT = r"[A-Za-z%]+(?:-[A-Za-z%]+)?"
_SUFFIX_RE = re.compile(rf"^(?P<core>.+?)-\s*(?P<qty>{_QTY})\s*(?P<unit>{_UNIT})\Z")

_LEADING_INCH_RE = re.compile(rf"^(?P<qty>{_NUM})\s*\\?\"\s*")

GLOBAL_LABEL_RULES = [
    (re.compile(r"^견적서?\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^부품\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^USP\s+"), "prefix"),  # \s+ 필수: 'USP-110-2107'(Monucla 자체 코드)처럼 공백 없이
                                          # 바로 하이픈/숫자가 오는 경우까지 잘못 걸리는 것을 방지
    (re.compile(r"^시리얼\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^일련\s*번호\s*[:：]?\s*"), "prefix"),
    (re.compile(r"^모델명\s*[:：]?\s*"), "prefix"),
    (re.compile(r"\s*외\s*\Z"), "suffix"),
]
GLOBAL_UI_RE = re.compile(r"UI\Z", re.IGNORECASE)

# 이미 완료된 상위60(상위20+차상위40) - 이번 범위에서 제외
TOP60_MFRS = {
    # 상위20 (1.41.51, v2.3)
    "Sartorius", "Thermo Fisher Scientific", "Sigma-Aldrich", "Sigma", "Merck Millipore",
    "Agilent", "대한과학", "삼전순약공업", "Corning", "USP", "Mettler Toledo", "Invitrogen",
    "Waters", "Cytiva", "유코", "Cell Signaling Technology", "Roche", "Eppendorf", "BD", "Gibco",
    # 차상위40 (1.41.52, v3.2)
    "Charles River", "Km", "R&D Systems", "Combi-Blocks", "Abcam", "PerkinElmer", "대정화금",
    "BioLegend", "SPL", "Saint-Gobain", "Supelco", "싸토리우스코리아", "영인에스티", "TCI", "비티알",
    "DURAN", "Promega", "DAE", "EP", "Cobetter", "MedChemExpress", "ProteinSimple", "GEMÜ",
    "Lonza", "Metrohm", "Avantor", "Beckman Coulter", "Agilent Technologies", "(주)BB", "TRC",
    "Repligen", "Millipore", "Isolab", "Siemens", "Daejung", "Gilson", "DUKSAN", "Rhawn",
    "시너지이노베이션", "Advanced Instruments",
}

_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _unit_matches(unit_up: str):
    if unit_up in COMPOUND_UNITS:
        return unit_up
    if "-" in unit_up:
        return None
    if unit_up in UNIT_WHITELIST:
        return unit_up
    for wl in UNIT_WHITELIST:
        if unit_up.startswith(wl):
            return wl
    return None


def compute_tobe(raw_pn: str):
    v = str(raw_pn).strip()
    removed_parts = []

    for pattern, kind in GLOBAL_LABEL_RULES:
        m = pattern.search(v) if kind == "suffix" else pattern.match(v)
        if m and m.group(0):
            removed_parts.append(m.group(0))
            v = pattern.sub("", v).strip()
            break

    if v.startswith("#"):
        removed_parts.append("#")
        v = v[1:].strip()

    if v.startswith("[") and v.endswith("]") and len(v) >= 2:
        removed_parts.append(v[0] + v[-1])
        v = v[1:-1].strip()

    m = GLOBAL_UI_RE.search(v)
    if m:
        removed_parts.append(v[m.start():])
        v = v[: m.start()]

    m = _LEADING_INCH_RE.match(v)
    if m:
        removed_parts.append(m.group(0))
        v = v[m.end():].strip()

    m = _SUFFIX_RE.match(v)
    if m:
        unit_up = m.group("unit").upper()
        matched = _unit_matches(unit_up)
        if matched:
            suffix_text = v[len(m.group("core")):]
            removed_parts.append(suffix_text)
            v = m.group("core").strip()

    separated_text = "".join(removed_parts)
    return v, separated_text


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        d1, col, d2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{d1}{get_column_letter(idx)}{d2}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)
    print(f"[복사] {SRC.name} -> {DST.name}")

    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb[SHEET]

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed
        if fixed != old_formula:
            print(f"[수식보정] {coord}: {old_formula} -> {fixed}")

    headers = ["품번_To-Be", "품번_To-Be_비고", "품번_To-Be_분리텍스트"]
    widths = [22, 32, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].width = w
        ws.cell(row=HEADER_ROW, column=INSERT_AT + i, value=h)

    df = pd.read_excel(SRC, sheet_name=SHEET, header=HEADER_ROW - 1)
    col_pn_name = df.columns[COL_PN - 1]
    col_mfr_name = df.columns[COL_MFR_CLEAN_OLD - 1]

    n_filled = 0
    n_sep = 0
    for i, row in df.iterrows():
        excel_row = i + DATA_START_ROW
        pn = row[col_pn_name]
        mfr = row[col_mfr_name]
        if pd.isna(pn):
            continue
        if mfr in TOP60_MFRS:
            continue  # 이미 완료된 상위60은 건드리지 않음
        tobe, sep = compute_tobe(pn)
        ws.cell(row=excel_row, column=INSERT_AT, value=tobe)
        n_filled += 1
        if sep:
            ws.cell(row=excel_row, column=INSERT_AT + 2, value=sep)
            n_sep += 1

    wb.save(DST)
    print(f"[요약] 품번_To-Be 채운 행: {n_filled}건")
    print(f"[요약] 분리텍스트 채운 행: {n_sep}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
