# -*- coding: utf-8 -*-
r"""
v1.6 -> v1.8

O열_2차판정 / O열_3차판정을 기존 값(2차 33건: 수작업 12건+21건, 3차 253건: 숫자0개 규칙)에서
아래처럼 명확하고 상호배타적인 기준으로 재정의해서 "같은 열에" 덮어쓴다(신규 열 추가 아님).

우선순위(1차판정이 [품번]인 행에 한해 순서대로 적용, 앞 단계에서 걸리면 뒷 단계는 안 봄):
  2차판정 = 품목명과 품번 값이 완전히 동일(공백 trim 후 exact match) -> [품명일치]
            숫자 포함 여부 상관없이 전수 적용 (기존엔 숫자 없는 것 위주로만 사람이 골랐었는데,
            "기준은 품명 일치 여부"라는 원칙을 그대로 전수 적용하기로 함 -> PD98059, CM-A99 등
            숫자 포함 값 9건도 포함됨)
  3차판정 = (2차에서 안 걸린 것 중) 값에 숫자가 하나도 없음 -> [비품번]
            (라벨 텍스트는 이후 step3에서 [숫자없음]으로 재변경됨. 여기선 아직 [비품번] 그대로 씀)

2차/3차가 바뀌면 "품번(잠정)" 후보 모수 자체가 달라지므로, O열_4차판정/검토메모(코드뭉침/설명문혼입/
단위값/구조식의심/슬래시검토)도 새 모수 기준으로 처음부터 다시 계산해서 반영한다(step1과 동일 규칙).

[폴더 이력] 1~3차판정 라벨 기준 자체를 다시 정리하는 작업이라 4차분리(1.41.44)가 아니라
1.41.43.5_품번_1~3차_작업_개편 폴더 소속으로 재배치됨. 진짜 4차판정은 이 작업이 끝난 v1.9를
v2.1로 승격해서 1.41.44에서 새로 시작함.
SRC는 1.41.43에서 완결된 v1.6을 그대로 입력으로 씀.

[버그 수정 이력] 최초 실행 시 `ws.cell(row, col, value=None)`으로 셀을 지우려 했으나, openpyxl에서
value=None은 "지우기"가 아니라 "값 안 주고 조회"로 처리되어(파라미터 기본값과 구분 불가) 2차판정
클리어가 무시되고 v1.6의 옛날 값이 25건 남는 버그가 있었음. `.value = None` 형태로 수정 후 재실행.
(다행히 4차판정/pool 결과 자체엔 영향 없었음 — 전수 검증 완료)
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43_품번_3차분리\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번3차판정_v1.6.xlsx")
DST = Path(r"C:\Users\정별\1_Work\1.41_동아쏘시오그룹(2)_데이터_표준화\1.41.43.5_품번_1~3차_작업_개편\20_결과\260810_S-TEPS_입고실적만 ◆_최근3개년_uniq_품번4차판정_v1.8.xlsx")

SHEET = "Steps_중복제거_32359"
HEADER_ROW = 4
DATA_START_ROW = 5

COL_품목명 = 13
COL_품번 = 15
COL_1차판정 = 16
COL_2차판정 = 17
COL_3차판정 = 18

INSERT_AT = 19  # O열_3차판정(18) 다음
N_NEW_COLS = 2  # O열_4차판정 + O열_4차_검토메모

LABEL_ITEM = "[품번]"
LABEL_NAME_MATCH = "[품명일치]"
LABEL_NO_DIGIT = "[비품번]"

LABEL_BUNDLE = "[코드뭉침]"
LABEL_DESC = "[설명문혼입]"
LABEL_UNIT = "[단위값]"
LABEL_STRUCT = "[구조식의심]"
NOTE_SLASH = "슬래시포함(수동검토필요)"

UNIT_ONLY_WHITELIST = {"50g", "25mg", "50mg", "500mg", "8mm"}
STRUCT_WHITELIST = {"OC1(CCOS(C2=CC=C(C)C=C2)(=O)=O)CCN(C(OCC3=CC=CC=C3"}

_FOR_RE = re.compile(r"\bfor\b", re.I)
_MFR_RE = re.compile(r"Mfr\s*#|Item\s*#", re.I)
_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_cols(formula: str, insert_at: int, n: int) -> str:
    def repl(m):
        dollar1, col, dollar2, row = m.groups()
        idx = column_index_from_string(col)
        if idx >= insert_at:
            idx += n
        return f"{dollar1}{get_column_letter(idx)}{dollar2}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def classify_4th(final_value: str) -> tuple[str, str]:
    v = final_value
    if "," in v:
        return LABEL_BUNDLE, ""
    if _FOR_RE.search(v) or _MFR_RE.search(v):
        return LABEL_DESC, ""
    if v in UNIT_ONLY_WHITELIST:
        return LABEL_UNIT, ""
    if v in STRUCT_WHITELIST:
        return LABEL_STRUCT, ""
    if "/" in v:
        return "", NOTE_SLASH
    return "", ""


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    n_name_match = 0
    n_no_digit = 0

    # ---- 1) 2차/3차판정 재계산 (기존 값 지우고 새 기준으로 덮어쓰기) ----
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_1차판정).value != LABEL_ITEM:
            continue

        ws.cell(row=r, column=COL_2차판정).value = None
        ws.cell(row=r, column=COL_3차판정).value = None

        raw = ws.cell(row=r, column=COL_품번).value
        name = ws.cell(row=r, column=COL_품목명).value
        s = "" if raw is None else str(raw).strip()
        n = "" if name is None else str(name).strip()

        if s and s == n:
            ws.cell(row=r, column=COL_2차판정, value=LABEL_NAME_MATCH)
            n_name_match += 1
            continue

        if s and not any(ch.isdigit() for ch in s):
            ws.cell(row=r, column=COL_3차판정, value=LABEL_NO_DIGIT)
            n_no_digit += 1

    # ---- 2) O열_4차판정/검토메모 열 삽입 + 새 모수 기준으로 재계산 ----
    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    ws.insert_cols(INSERT_AT, N_NEW_COLS)

    for coord, old_formula in formula_cells:
        fixed = shift_formula_cols(old_formula, INSERT_AT, N_NEW_COLS)
        ws[coord] = fixed

    for i in range(N_NEW_COLS):
        letter = get_column_letter(INSERT_AT + i)
        ws.column_dimensions[letter].hidden = False
        ws.column_dimensions[letter].width = 22

    ws.cell(row=HEADER_ROW, column=INSERT_AT, value="O열_4차판정")
    ws.cell(row=HEADER_ROW, column=INSERT_AT + 1, value="O열_4차_검토메모")

    # 품번_코어 열은 4차판정 삽입으로 한 칸 밀렸음
    COL_품번_코어_NEW = 19 + N_NEW_COLS  # 원래 19였던 O열_품번_코어

    counts4 = {LABEL_BUNDLE: 0, LABEL_DESC: 0, LABEL_UNIT: 0, LABEL_STRUCT: 0}
    n_note = 0
    n_pool = 0

    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_1차판정).value != LABEL_ITEM:
            continue
        if ws.cell(row=r, column=COL_2차판정).value or ws.cell(row=r, column=COL_3차판정).value:
            continue

        n_pool += 1
        core = ws.cell(row=r, column=COL_품번_코어_NEW).value
        raw = ws.cell(row=r, column=COL_품번).value
        final_value = str(core).strip() if core not in (None, "") else str(raw).strip()

        label, note = classify_4th(final_value)
        if label:
            ws.cell(row=r, column=INSERT_AT, value=label)
            counts4[label] += 1
        if note:
            ws.cell(row=r, column=INSERT_AT + 1, value=note)
            n_note += 1

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)

    print("[요약] 2차/3차판정 재계산 (품번(원문) vs 품목명 exact match 우선, 그 다음 숫자0개)")
    print(f"  O열_2차판정 [품명일치]: {n_name_match}건 (숫자 포함 값도 전수 포함)")
    print(f"  O열_3차판정 [비품번](숫자0개): {n_no_digit}건")
    print(f"  새 품번(잠정) 모수: {n_pool}건")
    print("[요약] O열_4차판정 재적용 결과 (새 모수 기준)")
    for k, c in counts4.items():
        print(f"  {k}: {c}건")
    print(f"  검토메모({NOTE_SLASH}): {n_note}건")
    print(f"[저장] {DST}")


if __name__ == "__main__":
    main()
